import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load criteria data (use v2 with improved extraction)
with open('outputs/criteria_graph_v2.json', 'r') as f:
    data = json.load(f)

total_students = data['metadata']['total_students']  # 36
criteria = data['criteria']

# Calculate total possible weight
max_weight = sum(c['weight'] for c in criteria.values())
print(f"Max possible weight: {max_weight:.2f}")

# Build student scores
student_criteria = {}  # student_id -> list of criteria keys they have
student_scores = {}

# Collect which criteria each student has
for crit_key, crit_data in criteria.items():
    for student in crit_data['students']:
        if student not in student_criteria:
            student_criteria[student] = []
        student_criteria[student].append(crit_key)

# Calculate scores for each student
for student, crit_list in student_criteria.items():
    # Base score
    raw_score = sum(criteria[c]['weight'] for c in crit_list)
    base_pct = (raw_score / max_weight) * 100

    # Rarity bonuses
    bonus = 0
    bonus_details = []
    for c in crit_list:
        prevalence = criteria[c]['count'] / total_students
        if prevalence <= 0.15:
            bonus += 1
            bonus_details.append((criteria[c]['display_name'], prevalence, 1))

    # Final score (capped at 100)
    final_score = min(100, base_pct + bonus)

    student_scores[student] = {
        'raw_score': raw_score,
        'base_pct': base_pct,
        'bonus': bonus,
        'final_score': final_score,
        'criteria_count': len(crit_list),
        'bonus_details': bonus_details
    }

# Sort by final score
ranked = sorted(student_scores.items(), key=lambda x: x[1]['final_score'], reverse=True)

# Print results
print("\n" + "="*80)
print("NEW GRADES WITH RARITY BONUS SYSTEM")
print("="*80)
print(f"{'Rank':<5} {'Student':<10} {'Base%':<8} {'Bonus':<6} {'Final':<8} {'Criteria':<10}")
print("-"*80)
total_criteria = len(criteria)
for rank, (student, scores) in enumerate(ranked, 1):
    print(f"{rank:<5} {student:<10} {scores['base_pct']:<8.1f} +{scores['bonus']:<5} {scores['final_score']:<8.1f} {scores['criteria_count']}/{total_criteria}")

# Create Excel file
wb = Workbook()
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
bonus_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
top_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Sheet 1: Grades
ws = wb.active
ws.title = "Grades"
headers = ["Rank", "Student ID", "Base Score %", "Rarity Bonus", "Final Grade", "Criteria Count", "Bonus Criteria"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for row, (student, scores) in enumerate(ranked, 2):
    bonus_names = ", ".join([f"{d[0]} (+{d[2]})" for d in scores['bonus_details']]) if scores['bonus_details'] else "None"
    values = [row-1, student, scores['base_pct'], scores['bonus'], scores['final_score'], f"{scores['criteria_count']}/{total_criteria}", bonus_names]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center') if col < 7 else Alignment(wrap_text=True)
        if col in [3, 5]:
            cell.number_format = '0.1'
    if scores['final_score'] >= 90:
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = top_fill

col_widths = [6, 12, 14, 12, 12, 14, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# Sheet 2: Rarity Criteria
ws2 = wb.create_sheet("Rarity Criteria")
headers2 = ["Criterion", "Prevalence", "Bonus", "Students Who Have It"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

rare_criteria = [(k, v) for k, v in criteria.items() if v['count'] / total_students <= 0.15]
rare_criteria.sort(key=lambda x: x[1]['count'])

for row, (key, crit) in enumerate(rare_criteria, 2):
    prev = crit['count'] / total_students
    bonus = 1
    ws2.cell(row=row, column=1, value=crit['display_name']).border = thin_border
    ws2.cell(row=row, column=2, value=f"{prev*100:.1f}%").border = thin_border
    ws2.cell(row=row, column=3, value=f"+{bonus}").border = thin_border
    ws2.cell(row=row, column=4, value=", ".join(crit['students'])).border = thin_border

for i, w in enumerate([40, 12, 8, 80], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save('outputs/grades.xlsx')
print(f"\nSaved: outputs/grades.xlsx")
