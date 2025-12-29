#!/usr/bin/env python3
"""Process WorkSubmissions04 and extract grade information from PDFs."""

import os
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def extract_info_from_pdf_text(pdf_text):
    """Extract the 5 required pieces of information from PDF text."""
    info = {
        'group_code': '',
        'student1': '',
        'student2': '',
        'github_link': '',
        'suggested_grade': ''
    }

    # Patterns for extraction
    patterns = {
        'group_code': [
            r'1\.\s*(?:Group [Cc]ode|Codegroupe|קוד קבוצה):\s*([^\n]+)',
            r'Group [Cc]ode:\s*([^\n]+)',
            r'קוד קבוצה:\s*([^\n]+)',
        ],
        'student1': [
            r'2\.\s*(?:Student one|Member A|שם חבר צוות ראשון):\s*([^\n]+)',
            r'Student one:\s*([^\n]+)',
            r'Member A:\s*([^\n]+)',
        ],
        'student2': [
            r'3\.\s*(?:Student two|Member B|שם חבר צוות שני):\s*([^\n]+)',
            r'Student two:\s*([^\n]+)',
            r'Member B:\s*([^\n]+)',
        ],
        'github_link': [
            r'4\.\s*(?:Repo link|GitHub Repository|קישור לריפו):\s*([^\n]+)',
            r'GitHub Repository:\s*([^\n]+)',
            r'Repo link:\s*([^\n]+)',
            r'(https://github\.com/[^\s\)]+)',
        ],
        'suggested_grade': [
            r'5\.\s*(?:Grade suggestion|הציון העצמי שלי):\s*(\d+)',
            r'Grade suggestion:\s*(\d+)',
            r'הציון העצמי שלי\s*[:]?\s*(\d+)',
        ]
    }

    # Try each pattern for each field
    for field, pattern_list in patterns.items():
        for pattern in pattern_list:
            match = re.search(pattern, pdf_text, re.IGNORECASE)
            if match:
                info[field] = match.group(1).strip()
                break

    # Clean up GitHub URL
    github_match = re.search(r'https://github\.com/[^\s\)]+', info['github_link'])
    if github_match:
        info['github_link'] = github_match.group(0)

    # Extract just the number from suggested grade
    grade_match = re.search(r'(\d+)', info['suggested_grade'])
    if grade_match:
        info['suggested_grade'] = grade_match.group(1)

    return info


def create_excel_workbook(output_file):
    """Create a new Excel workbook with headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Submissions"

    headers = [
        "Participant ID",
        "Group Code",
        "Student 1",
        "Student 2",
        "GitHub Repository",
        "Suggested Grade",
        "PDF Filename"
    ]

    # Write headers with formatting
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 50

    wb.save(output_file)
    return wb


def get_participant_folders():
    """Get all participant folders from WorkSubmissions04."""
    base_path = Path('WorkSubmissions04')
    folders = sorted([
        d for d in base_path.iterdir()
        if d.is_dir() and d.name.startswith('Participant_')
    ])
    return folders


def find_main_pdf(folder_path):
    """Find the main assignment PDF (not submission_keys.pdf)."""
    pdf_files = list(folder_path.glob('*.pdf'))
    # Filter out submission_keys.pdf
    main_pdfs = [p for p in pdf_files if 'submission_keys' not in p.name.lower()]
    return main_pdfs[0] if main_pdfs else (pdf_files[0] if pdf_files else None)


def extract_participant_id(folder_name):
    """Extract participant ID from folder name."""
    match = re.search(r'Participant_(\d+)', folder_name)
    return match.group(1) if match else folder_name


# This will be populated by the agent reading PDFs
submissions_data = []

if __name__ == '__main__':
    folders = get_participant_folders()
    print(f"Found {len(folders)} participant folders")

    for folder in folders:
        participant_id = extract_participant_id(folder.name)
        pdf_file = find_main_pdf(folder)

        if pdf_file:
            print(f"Participant {participant_id}: {pdf_file.name}")
            submissions_data.append({
                'participant_id': participant_id,
                'folder': folder.name,
                'pdf_path': str(pdf_file),
                'pdf_filename': pdf_file.name
            })
        else:
            print(f"WARNING: No PDF found for Participant {participant_id}")
