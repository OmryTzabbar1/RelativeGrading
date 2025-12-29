#!/usr/bin/env python3
"""
Complete evaluation workflow for WorkSubmissions05
Following EXTRACTION.md and CATEGORIES.md rules
"""

import os
import json
import re
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Base paths
BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

# Student data structure
student_criteria = defaultdict(set)  # {student_id: {criterion1, criterion2, ...}}
criterion_sources = defaultdict(lambda: defaultdict(list))  # {student_id: {criterion: [file1, file2]}}

# Criteria normalization map
NORMALIZATION_MAP = {
    # Testing
    "unit test": "Unit_Tests",
    "unit tests": "Unit_Tests",
    "unit testing": "Unit_Tests",
    "integration test": "Integration_Tests",
    "integration tests": "Integration_Tests",
    "integration testing": "Integration_Tests",
    "e2e test": "E2E_Tests",
    "e2e tests": "E2E_Tests",
    "end-to-end test": "E2E_Tests",
    "end-to-end tests": "E2E_Tests",
    "test coverage": "Test_Coverage_Metrics",
    "code coverage": "Test_Coverage_Metrics",
    "coverage": "Test_Coverage_Metrics",

    # Documentation
    "readme": "README",
    "readme.md": "README",
    "api documentation": "API_Documentation",
    "api docs": "API_Documentation",
    "api reference": "API_Documentation",
    "changelog": "Changelog",
    "contributing guide": "Contributing_Guide",
    "contributing": "Contributing_Guide",
    "installation instructions": "Installation_Instructions",
    "setup guide": "Installation_Instructions",
    "usage guide": "Usage_Guide",
    "troubleshooting guide": "Troubleshooting_Guide",
    "quick start": "Quick_Start_Guide",
    "quickstart": "Quick_Start_Guide",

    # Planning
    "prd": "PRD_Document",
    "prd document": "PRD_Document",
    "product requirements": "PRD_Document",
    "architecture": "Architecture_Documentation",
    "architecture documentation": "Architecture_Documentation",
    "problem statement": "Problem_Statement",
    "solution overview": "Solution_Overview",
    "proposed solution": "Solution_Overview",
    "use case": "Use_Case_Documentation",
    "use cases": "Use_Case_Documentation",
    "functional requirements": "Functional_Requirements",
    "non-functional requirements": "Non_Functional_Requirements",
    "success metrics": "Success_Metrics",
    "kpis": "Success_Metrics",
    "assumptions": "Assumptions_Documentation",
    "constraints": "Constraints_Documentation",
    "limitations": "Constraints_Documentation",
    "cost analysis": "Cost_Analysis",
    "risk analysis": "Risk_Analysis",
    "roadmap": "Roadmap",
    "user stories": "User_Stories",
    "design decisions": "Design_Decision_Documentation",
    "trade-offs": "Trade_offs_Analysis",

    # DevOps
    "ci/cd": "CI_CD_Pipeline",
    "ci cd": "CI_CD_Pipeline",
    "continuous integration": "CI_CD_Pipeline",
    "docker": "Docker",
    "dockerfile": "Docker",
    "containerization": "Docker",
    "github actions": "GitHub_Actions",

    # Quality
    "eslint": "ESLint_Configuration",
    "pylint": "Pylint_Configuration",
    "prettier": "Prettier_Formatting",
    "black": "Black_Formatting",
    "pre-commit": "Pre_commit_Hooks",
    "pre-commit hooks": "Pre_commit_Hooks",
    "type checking": "Type_Checking",
    "typescript": "TypeScript_Type_Checking",
    "mypy": "Mypy_Type_Checking",
    "code review": "Code_Review_Process",
    "code style guide": "Code_Style_Guide",
    "pep 8": "PEP8_Compliance",
    "pep8": "PEP8_Compliance",

    # Visuals
    "screenshots": "Screenshots",
    "screenshot": "Screenshots",
    "architecture diagram": "Architecture_Diagram",
    "demo video": "Demo_Video",
}

# File-based criteria detection
FILENAME_CRITERIA = {
    "prd.md": "PRD_Document",
    "productRequirements.md": "PRD_Document",
    "testing.md": "Testing_Documentation",
    "test.md": "Testing_Documentation",
    "contributing.md": "Contributing_Guide",
    "quickstart.md": "Quick_Start_Guide",
    "quick-start.md": "Quick_Start_Guide",
    "architecture.md": "Architecture_Documentation",
    "design.md": "Architecture_Documentation",
    "changelog.md": "Changelog",
    "changes.md": "Changelog",
    "api.md": "API_Documentation",
    "api_docs.md": "API_Documentation",
    "roadmap.md": "Roadmap",
    "deployment.md": "Deployment_Guide",
    "deploy.md": "Deployment_Guide",
    "troubleshooting.md": "Troubleshooting_Guide",
    "faq.md": "Troubleshooting_Guide",
}

# Categories from CATEGORIES.md
CATEGORIES = {
    "Documentation": [
        "readme", "api", "user guide", "usage guide", "changelog", "contributing",
        "installation", "setup", "getting started", "faq", "manual", "reference",
        "troubleshooting", "quick start"
    ],
    "Planning": [
        "prd", "product requirements", "architecture", "design document", "technical spec",
        "roadmap", "milestones", "requirements", "system design", "wireframes", "mockups",
        "user stories", "use case", "problem statement", "solution", "objectives",
        "success metrics", "kpi", "assumptions", "constraints", "limitations", "scope",
        "design decision", "rationale", "trade-off", "alternatives", "technology",
        "user persona", "user flow", "timeline", "schedule", "project plan"
    ],
    "Testing": [
        "unit test", "integration test", "e2e", "end-to-end", "test coverage",
        "code coverage", "pytest", "jest", "mocha", "junit", "test suite",
        "automated test", "regression", "smoke test", "load test", "performance test",
        "test case", "tdd"
    ],
    "DevOps": [
        "ci/cd", "ci cd", "continuous integration", "continuous deployment",
        "github actions", "gitlab ci", "jenkins", "docker", "kubernetes", "k8s",
        "deployment", "deploy", "aws", "azure", "gcp", "cloud", "terraform",
        "monitoring", "logging", "nginx", "apache", "ssl", "https", "environment",
        "build pipeline", "release"
    ],
    "Research": [
        "research", "analysis", "data analysis", "jupyter", "notebook",
        "data exploration", "experiment", "hypothesis", "findings", "results",
        "literature review", "benchmark", "comparison", "survey", "user research",
        "insights", "statistical", "machine learning", "model evaluation"
    ],
    "Visuals": [
        "screenshot", "diagram", "flowchart", "chart", "graph", "demo video",
        "gif", "animated", "ui preview", "before/after", "sequence diagram",
        "erd", "entity relationship", "class diagram", "infographic"
    ],
    "CodeQuality": [
        "linting", "linter", "eslint", "pylint", "ruff", "flake8", "prettier",
        "formatting", "black", "autopep8", "yapf", "type checking", "typescript",
        "mypy", "pyright", "code review", "peer review", "pre-commit", "git hooks",
        "husky", "lint-staged", "refactoring", "clean code", "solid", "design pattern",
        "code style", "style guide", "pep 8", "pep8", "airbnb", "google style",
        "docstring", "static analysis", "sonarqube", "quality gate", "quality check",
        "setup.py", "pyproject.toml", "package.json", "dependency", "venv",
        "virtual environment"
    ],
    "Business": [
        "cost analysis", "cost breakdown", "budget", "roi", "return on investment",
        "market research", "market analysis", "customer persona", "business case",
        "business model", "pricing", "monetization", "competitive analysis",
        "competitor", "swot", "value proposition", "target audience", "stakeholder",
        "revenue", "go-to-market", "risk analysis", "risk assessment", "risk mitigation"
    ]
}

# Positive indicators for implementation
POSITIVE_INDICATORS = [
    "we built", "we created", "we developed", "we implemented",
    "project includes", "this project includes", "includes",
    "features:", "key features:", "implemented", "created", "developed",
    "built", "added", "integrated", "incorporated", "supports", "provides",
    "offers", "enables", "has", "contains", "comes with", "complete",
    "comprehensive", "full", "available:", "featuring:", "with support for",
    "capable of", "total tests:", "test cases:", "coverage:", "version"
]

# Negative indicators (don't count these)
NEGATIVE_INDICATORS = [
    "todo:", "fixme:", "will add", "planning to", "planned for", "future work:",
    "next steps:", "not yet implemented", "not complete", "missing", "out of scope",
    "skipped", "omitted", "no ", "not working", "might be", "considering",
    "if time permits", "would be nice", "ideally", "optional"
]

def normalize_criterion(text):
    """Normalize criterion name"""
    text_lower = text.lower().strip()

    # Check normalization map
    if text_lower in NORMALIZATION_MAP:
        return NORMALIZATION_MAP[text_lower]

    # Convert to Title_Case_Format
    words = re.split(r'[\s\-_/]+', text)
    normalized = '_'.join(word.capitalize() for word in words if word)
    return normalized

def is_negative_context(line):
    """Check if line contains negative indicators"""
    line_lower = line.lower()
    return any(neg in line_lower for neg in NEGATIVE_INDICATORS)

def is_positive_context(line):
    """Check if line contains positive indicators"""
    line_lower = line.lower()
    return any(pos in line_lower for pos in POSITIVE_INDICATORS)

def extract_from_filename(filename):
    """Extract criteria based on filename"""
    filename_lower = filename.lower()
    criteria = []

    for pattern, criterion in FILENAME_CRITERIA.items():
        if pattern in filename_lower:
            criteria.append(criterion)

    return criteria

def extract_from_headers(content, filename):
    """Extract criteria from section headers"""
    criteria = []
    lines = content.split('\n')

    header_patterns = {
        r'^#+\s*(testing|tests)': 'Testing_Documentation',
        r'^#+\s*unit\s*tests?': 'Unit_Tests',
        r'^#+\s*installation|setup|getting started': 'Installation_Instructions',
        r'^#+\s*usage|how to use': 'Usage_Guide',
        r'^#+\s*features?': None,  # Scan features section
        r'^#+\s*screenshots?|demo|examples': 'Screenshots',
        r'^#+\s*architecture|design': 'Architecture_Documentation',
        r'^#+\s*ci/?cd|deployment|devops': None,  # Scan section
        r'^#+\s*api|endpoints|api reference': 'API_Documentation',
        r'^#+\s*contributing|development': 'Contributing_Guide',
        r'^#+\s*problem\s*statement|problem': 'Problem_Statement',
        r'^#+\s*solution|proposed solution': 'Solution_Overview',
        r'^#+\s*requirements|functional requirements': 'Functional_Requirements',
        r'^#+\s*use\s*cases?': 'Use_Case_Documentation',
        r'^#+\s*cost\s*analysis|budget|costs': 'Cost_Analysis',
        r'^#+\s*assumptions|constraints': 'Assumptions_Documentation',
        r'^#+\s*success\s*metrics|kpis|metrics': 'Success_Metrics',
        r'^#+\s*risk\s*analysis|risks': 'Risk_Analysis',
        r'^#+\s*roadmap|future work|next steps': 'Roadmap',
    }

    for i, line in enumerate(lines):
        for pattern, criterion in header_patterns.items():
            if re.match(pattern, line, re.IGNORECASE):
                if criterion:
                    # Check if section has content (not TODO)
                    next_content = '\n'.join(lines[i+1:i+10])
                    if not is_negative_context(next_content):
                        criteria.append(criterion)

    return criteria

def extract_from_content(content, filename):
    """Extract criteria from markdown content"""
    criteria = []
    lines = content.split('\n')

    # Quality standards
    quality_patterns = {
        r'\beslint\b': 'ESLint_Configuration',
        r'\bpylint\b': 'Pylint_Configuration',
        r'\bprettier\b': 'Prettier_Formatting',
        r'\bblack\s+(formatter|formatting)\b': 'Black_Formatting',
        r'\bpre-?commit\s*(hooks?|configuration)\b': 'Pre_commit_Hooks',
        r'\btypescript\b.*\b(strict|type checking)\b': 'TypeScript_Type_Checking',
        r'\bmypy\b': 'Mypy_Type_Checking',
        r'\bcode\s*review\b': 'Code_Review_Process',
        r'\bcode\s*style\s*guide\b': 'Code_Style_Guide',
        r'\bpep\s*8\b': 'PEP8_Compliance',
    }

    # Testing patterns
    test_patterns = {
        r'\bunit\s*tests?\b': 'Unit_Tests',
        r'\bintegration\s*tests?\b': 'Integration_Tests',
        r'\be2e\s*tests?\b|\bend-to-end\s*tests?\b': 'E2E_Tests',
        r'\b(\d+)%\s*(code\s*)?coverage\b': 'Test_Coverage_Metrics',
        r'\btotal\s*tests?:\s*[~]?(\d+)': 'Unit_Tests',
    }

    # Planning patterns
    planning_patterns = {
        r'\bproblem\s*statement\b': 'Problem_Statement',
        r'\bsolution\s*overview\b|\bproposed\s*solution\b': 'Solution_Overview',
        r'\bsuccess\s*metrics\b|\bkpis\b': 'Success_Metrics',
        r'\buse\s*cases?\b': 'Use_Case_Documentation',
        r'\bcost\s*analysis\b': 'Cost_Analysis',
        r'\bassumptions?\b': 'Assumptions_Documentation',
        r'\bconstraints?\b|\blimitations?\b': 'Constraints_Documentation',
        r'\brisk\s*analysis\b': 'Risk_Analysis',
        r'\bfunctional\s*requirements?\b': 'Functional_Requirements',
    }

    all_patterns = {**quality_patterns, **test_patterns, **planning_patterns}

    for line in lines:
        # Skip negative context
        if is_negative_context(line):
            continue

        # Check for positive context
        if is_positive_context(line):
            for pattern, criterion in all_patterns.items():
                if re.search(pattern, line, re.IGNORECASE):
                    criteria.append(criterion)

    return criteria

def process_student(student_dir):
    """Process all markdown files for a student"""
    student_id = student_dir.name
    print(f"\nProcessing {student_id}...")

    # Find all markdown files
    md_files = list(student_dir.rglob("*.md"))
    print(f"  Found {len(md_files)} markdown files")

    for md_file in md_files:
        print(f"    Reading {md_file.relative_to(student_dir)}")

        try:
            content = md_file.read_text(encoding='utf-8', errors='ignore')
        except Exception as e:
            print(f"    ERROR reading {md_file}: {e}")
            continue

        # Extract from filename
        filename_criteria = extract_from_filename(md_file.name)
        for criterion in filename_criteria:
            student_criteria[student_id].add(criterion)
            criterion_sources[student_id][criterion].append(str(md_file.relative_to(student_dir)))

        # Extract from headers
        header_criteria = extract_from_headers(content, md_file.name)
        for criterion in header_criteria:
            student_criteria[student_id].add(criterion)
            criterion_sources[student_id][criterion].append(str(md_file.relative_to(student_dir)))

        # Extract from content
        content_criteria = extract_from_content(content, md_file.name)
        for criterion in content_criteria:
            student_criteria[student_id].add(criterion)
            criterion_sources[student_id][criterion].append(str(md_file.relative_to(student_dir)))

    print(f"  Total criteria found: {len(student_criteria[student_id])}")

def categorize_criterion(criterion_name):
    """Categorize a criterion based on keywords"""
    criterion_lower = criterion_name.lower().replace('_', ' ')

    for category, keywords in CATEGORIES.items():
        for keyword in keywords:
            if keyword in criterion_lower:
                return category

    return "Uncategorized"

def build_criteria_graph():
    """Build the criteria graph structure"""
    criteria_graph = {
        "metadata": {
            "total_students": len(student_criteria),
            "evaluation_date": datetime.now().isoformat(),
            "total_criteria": 0
        },
        "criteria": {}
    }

    # Aggregate all criteria
    all_criteria = set()
    for student_id, criteria_set in student_criteria.items():
        all_criteria.update(criteria_set)

    criteria_graph["metadata"]["total_criteria"] = len(all_criteria)

    # Build criteria objects
    for criterion in sorted(all_criteria):
        students_with_criterion = [
            student_id for student_id, criteria_set in student_criteria.items()
            if criterion in criteria_set
        ]

        count = len(students_with_criterion)
        weight = count / len(student_criteria)

        # Apply rarity bonus (≤15% prevalence)
        rarity_bonus = 1.0 if weight <= 0.15 else 0.0

        category = categorize_criterion(criterion)

        criteria_graph["criteria"][criterion] = {
            "display_name": criterion.replace('_', ' '),
            "students": students_with_criterion,
            "count": count,
            "weight": round(weight, 4),
            "rarity_bonus": rarity_bonus,
            "category": category
        }

    return criteria_graph

def calculate_scores(criteria_graph):
    """Calculate scores and grades for all students"""
    student_scores = {}

    # Calculate max possible score
    max_possible = sum(
        c["weight"] + c["rarity_bonus"]
        for c in criteria_graph["criteria"].values()
    )

    for student_id, criteria_set in student_criteria.items():
        raw_score = 0.0

        for criterion in criteria_set:
            if criterion in criteria_graph["criteria"]:
                raw_score += criteria_graph["criteria"][criterion]["weight"]
                raw_score += criteria_graph["criteria"][criterion]["rarity_bonus"]

        percentage = (raw_score / max_possible) * 100 if max_possible > 0 else 0

        student_scores[student_id] = {
            "raw_score": round(raw_score, 4),
            "max_possible": round(max_possible, 4),
            "percentage": round(percentage, 2),
            "criteria_count": len(criteria_set),
            "criteria": list(criteria_set)
        }

    # Calculate relative grades
    best_percentage = max(s["percentage"] for s in student_scores.values())

    for student_id in student_scores:
        relative_grade = (student_scores[student_id]["percentage"] / best_percentage) * 100
        student_scores[student_id]["relative_grade"] = round(relative_grade, 2)

    # Assign ranks
    sorted_students = sorted(
        student_scores.items(),
        key=lambda x: x[1]["percentage"],
        reverse=True
    )

    for rank, (student_id, score_data) in enumerate(sorted_students, 1):
        student_scores[student_id]["rank"] = rank

    return student_scores

def generate_excel_outputs(criteria_graph, student_scores):
    """Generate Excel output files"""

    # 1. grades.xlsx - Simple grades table
    wb_grades = openpyxl.Workbook()
    ws = wb_grades.active
    ws.title = "Grades"

    headers = ["Student ID", "Raw Score", "Max Possible", "Percentage", "Grade", "Rank", "Criteria Count"]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for student_id, scores in sorted(student_scores.items(), key=lambda x: x[1]["rank"]):
        ws.append([
            student_id,
            scores["raw_score"],
            scores["max_possible"],
            scores["percentage"],
            scores["relative_grade"],
            scores["rank"],
            scores["criteria_count"]
        ])

    grades_file = OUTPUT_DIR / "grades.xlsx"
    wb_grades.save(grades_file)
    print(f"\nSaved: {grades_file}")

    # 2. Student_Evaluation_Report.xlsx - Comprehensive multi-sheet report
    wb_report = openpyxl.Workbook()
    wb_report.remove(wb_report.active)

    # Sheet 1: Summary
    ws_summary = wb_report.create_sheet("Summary")
    ws_summary.append(["WorkSubmissions05 Evaluation Report"])
    ws_summary.append(["Evaluation Date", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_summary.append(["Total Students", len(student_criteria)])
    ws_summary.append(["Total Criteria", len(criteria_graph["criteria"])])
    ws_summary.append([])
    ws_summary.append(["Category", "Criteria Count"])

    # Category breakdown
    category_counts = defaultdict(int)
    for criterion, data in criteria_graph["criteria"].items():
        category_counts[data["category"]] += 1

    for category, count in sorted(category_counts.items()):
        ws_summary.append([category, count])

    # Sheet 2: Student Grades
    ws_grades = wb_report.create_sheet("Student Grades")
    ws_grades.append(headers)

    for cell in ws_grades[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for student_id, scores in sorted(student_scores.items(), key=lambda x: x[1]["rank"]):
        ws_grades.append([
            student_id,
            scores["raw_score"],
            scores["max_possible"],
            scores["percentage"],
            scores["relative_grade"],
            scores["rank"],
            scores["criteria_count"]
        ])

    # Sheet 3: Criteria Details
    ws_criteria = wb_report.create_sheet("Criteria Details")
    ws_criteria.append(["Criterion", "Category", "Count", "Weight", "Rarity Bonus", "Prevalence %"])

    for cell in ws_criteria[1]:
        cell.font = Font(bold=True)

    for criterion, data in sorted(criteria_graph["criteria"].items()):
        ws_criteria.append([
            data["display_name"],
            data["category"],
            data["count"],
            data["weight"],
            data["rarity_bonus"],
            round(data["weight"] * 100, 2)
        ])

    report_file = OUTPUT_DIR / "Student_Evaluation_Report.xlsx"
    wb_report.save(report_file)
    print(f"Saved: {report_file}")

def generate_markdown_report(criteria_graph, student_scores):
    """Generate EVALUATION_SUMMARY.md"""

    lines = []
    lines.append("# WorkSubmissions05 Evaluation Summary\n")
    lines.append(f"**Evaluation Date:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
    lines.append(f"**Total Students:** {len(student_criteria)}\n")
    lines.append(f"**Total Criteria Discovered:** {len(criteria_graph['criteria'])}\n")
    lines.append("")

    # Category breakdown
    category_counts = defaultdict(int)
    for criterion, data in criteria_graph["criteria"].items():
        category_counts[data["category"]] += 1

    lines.append("## Criteria by Category\n")
    for category in sorted(category_counts.keys()):
        count = category_counts[category]
        lines.append(f"- **{category}:** {count} criteria")
    lines.append("")

    # Top students
    lines.append("## Top 10 Students\n")
    top_students = sorted(student_scores.items(), key=lambda x: x[1]["rank"])[:10]

    for student_id, scores in top_students:
        lines.append(f"### {scores['rank']}. {student_id}")
        lines.append(f"- **Grade:** {scores['relative_grade']}/100")
        lines.append(f"- **Percentage:** {scores['percentage']}%")
        lines.append(f"- **Criteria Count:** {scores['criteria_count']}")
        lines.append("")

    # Grade distribution
    lines.append("## Grade Distribution\n")
    grade_ranges = {
        "90-100": 0,
        "80-89": 0,
        "70-79": 0,
        "60-69": 0,
        "Below 60": 0
    }

    for scores in student_scores.values():
        grade = scores["relative_grade"]
        if grade >= 90:
            grade_ranges["90-100"] += 1
        elif grade >= 80:
            grade_ranges["80-89"] += 1
        elif grade >= 70:
            grade_ranges["70-79"] += 1
        elif grade >= 60:
            grade_ranges["60-69"] += 1
        else:
            grade_ranges["Below 60"] += 1

    for range_name, count in grade_ranges.items():
        lines.append(f"- **{range_name}:** {count} students")
    lines.append("")

    # CodeQuality criteria
    cq_criteria = [
        (name, data) for name, data in criteria_graph["criteria"].items()
        if data["category"] == "CodeQuality"
    ]

    lines.append(f"## Code Quality Criteria ({len(cq_criteria)} total)\n")
    for criterion, data in sorted(cq_criteria, key=lambda x: x[1]["count"], reverse=True):
        lines.append(f"- **{data['display_name']}**: {data['count']} students ({data['weight']*100:.1f}%)")
    lines.append("")

    # Write file
    report_file = OUTPUT_DIR / "EVALUATION_SUMMARY.md"
    report_file.write_text('\n'.join(lines), encoding='utf-8')
    print(f"Saved: {report_file}")

def main():
    """Main evaluation workflow"""
    print("=" * 80)
    print("WorkSubmissions05 Evaluation - Complete Workflow")
    print("=" * 80)

    # Find all student directories
    student_dirs = sorted([
        d for d in BASE_DIR.iterdir()
        if d.is_dir() and d.name.startswith("Participant_")
    ])

    print(f"\nFound {len(student_dirs)} student folders")

    # STEP 3-4: Extract criteria from all students
    print("\n" + "=" * 80)
    print("STEP 3-4: Reading and Extracting Criteria")
    print("=" * 80)

    for student_dir in student_dirs:
        process_student(student_dir)

    # STEP 5: Build criteria graph
    print("\n" + "=" * 80)
    print("STEP 5: Building Criteria Graph")
    print("=" * 80)

    criteria_graph = build_criteria_graph()
    print(f"Total unique criteria: {len(criteria_graph['criteria'])}")

    # Save criteria graph
    graph_file = OUTPUT_DIR / "criteria_graph_final.json"
    with open(graph_file, 'w', encoding='utf-8') as f:
        json.dump(criteria_graph, f, indent=2)
    print(f"Saved: {graph_file}")

    # STEP 6: Categorize (already done in build_criteria_graph)
    print("\n" + "=" * 80)
    print("STEP 6: Categorization Complete")
    print("=" * 80)

    category_counts = defaultdict(int)
    for criterion, data in criteria_graph["criteria"].items():
        category_counts[data["category"]] += 1

    for category, count in sorted(category_counts.items()):
        print(f"  {category}: {count} criteria")

    # STEP 7: Calculate weights (already done in build_criteria_graph)
    # STEP 8: Calculate scores and grades
    print("\n" + "=" * 80)
    print("STEP 8: Calculating Scores and Grades")
    print("=" * 80)

    student_scores = calculate_scores(criteria_graph)

    # STEP 9: Generate outputs
    print("\n" + "=" * 80)
    print("STEP 9: Generating Output Files")
    print("=" * 80)

    generate_excel_outputs(criteria_graph, student_scores)
    generate_markdown_report(criteria_graph, student_scores)

    # Final summary
    print("\n" + "=" * 80)
    print("EVALUATION COMPLETE")
    print("=" * 80)
    print(f"Total Students: {len(student_criteria)}")
    print(f"Total Criteria: {len(criteria_graph['criteria'])}")
    print(f"CodeQuality Criteria: {category_counts.get('CodeQuality', 0)}")
    print("\nTop 3 Students:")

    top_3 = sorted(student_scores.items(), key=lambda x: x[1]["rank"])[:3]
    for student_id, scores in top_3:
        print(f"  {scores['rank']}. {student_id}")
        print(f"     Grade: {scores['relative_grade']}/100 | Criteria: {scores['criteria_count']}")

    print("\nOutputs saved to:", OUTPUT_DIR)

if __name__ == "__main__":
    main()
