#!/usr/bin/env python3
"""
Create individual Excel files in each student's submission folder
This script reads each student's PDF and creates a submission_info.xlsx file in their folder
"""

import os
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_student_excel(folder_path, participant_id, group_code, student1, student2, github, grade, pdf_filename):
    """Create an individual Excel file for a student submission."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission Info"

    # Headers
    headers = [
        "Field",
        "Value"
    ]

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    data = [
        ("Participant ID", participant_id),
        ("Group Code", group_code),
        ("Student 1", student1),
        ("Student 2", student2),
        ("GitHub Repository", github),
        ("Suggested Grade", grade),
        ("PDF Filename", pdf_filename)
    ]

    for row_idx, (field, value) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        cell = ws.cell(row=row_idx, column=2, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Make GitHub link clickable
        if field == "GitHub Repository" and value and value.startswith('http'):
            cell.hyperlink = value
            cell.font = Font(color="0563C1", underline="single")

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    # Save the Excel file in the student's folder
    output_path = folder_path / "submission_info.xlsx"
    wb.save(output_path)
    return output_path


def extract_participant_id(folder_name):
    """Extract participant ID from folder name."""
    match = re.search(r'Participant_(\d+)', folder_name)
    return match.group(1) if match else folder_name


def main():
    """Process all participant folders and create individual Excel files."""
    base_path = Path('.')
    participant_folders = sorted([
        d for d in base_path.iterdir()
        if d.is_dir() and d.name.startswith('Participant_')
    ])

    print(f"Found {len(participant_folders)} participant folders")
    print("This script expects Claude Code to extract the data from PDFs")
    print("and call create_student_excel() for each participant.\n")

    return participant_folders


if __name__ == '__main__':
    folders = main()
    for folder in folders:
        participant_id = extract_participant_id(folder.name)
        print(f"Participant {participant_id}: {folder.name}")
