#!/usr/bin/env python3
"""
Grade Comparison Tool
Compares student-project-evaluator grades with actual grades from PDF files.
"""

import sys
import os
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import ScatterChart, Reference, BarChart
from openpyxl.chart.series import DataPoint
import json

# Try to import PDF parsing libraries
try:
    import pdfplumber
    PDF_LIBRARY = 'pdfplumber'
except ImportError:
    try:
        import PyPDF2
        PDF_LIBRARY = 'PyPDF2'
    except ImportError:
        print("Error: No PDF library found. Install with: pip install pdfplumber")
        sys.exit(1)


def extract_grade_from_pdf(pdf_path):
    """Extract grade from PDF using available library"""

    if PDF_LIBRARY == 'pdfplumber':
        return extract_grade_pdfplumber(pdf_path)
    else:
        return extract_grade_pypdf2(pdf_path)


def extract_grade_pdfplumber(pdf_path):
    """Extract grade from PDF using pdfplumber"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            # Read all pages
            for page in pdf.pages:
                text += page.extract_text() or ""

            return parse_grade_from_text(text)
    except Exception as e:
        print(f"  Error reading PDF: {e}")
        return None


def extract_grade_pypdf2(pdf_path):
    """Extract grade from PDF using PyPDF2"""
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ""
            # Read all pages
            for page in reader.pages:
                text += page.extract_text() or ""

            return parse_grade_from_text(text)
    except Exception as e:
        print(f"  Error reading PDF: {e}")
        return None


def parse_grade_from_text(text):
    """Parse grade from PDF text content"""

    # Common patterns for grade extraction
    patterns = [
        r'Total\s+Grade[:\s]+(\d+(?:\.\d+)?)',
        r'Final\s+Grade[:\s]+(\d+(?:\.\d+)?)',
        r'Overall\s+Grade[:\s]+(\d+(?:\.\d+)?)',
        r'Grade[:\s]+(\d+(?:\.\d+)?)\s*%',
        r'Grade[:\s]+(\d+(?:\.\d+)?)\s*/\s*100',
        r'Score[:\s]+(\d+(?:\.\d+)?)',
        r'Total[:\s]+(\d+(?:\.\d+)?)\s*%',
        r'Final\s+Score[:\s]+(\d+(?:\.\d+)?)',
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            grade = float(match.group(1))
            # Validate grade is in reasonable range
            if 0 <= grade <= 100:
                return grade

    return None


def load_evaluator_grades(grades_path):
    """Load grades from evaluator's grades.xlsx"""

    print(f"Loading evaluator grades from: {grades_path}")

    wb = openpyxl.load_workbook(grades_path)
    ws = wb.active

    grades = {}

    # Skip header row
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        rank = row[0]
        student_id = str(row[1])
        final_grade = row[2]
        base_pct = row[3]
        bonus = row[4]
        criteria_count = row[5] if len(row) > 5 else None

        grades[student_id] = {
            'rank': rank,
            'final_grade': final_grade,
            'base_pct': base_pct,
            'bonus': bonus,
            'criteria_count': criteria_count
        }

    print(f"Found {len(grades)} students in evaluator results\n")
    return grades


def find_student_folders(worksubmissions_path):
    """Find all student folders in WorkSubmissions directory"""

    print(f"Scanning WorkSubmissions folder: {worksubmissions_path}")

    path = Path(worksubmissions_path)
    if not path.exists():
        raise FileNotFoundError(f"WorkSubmissions path does not exist: {worksubmissions_path}")

    folders = {}

    for folder in path.iterdir():
        if not folder.is_dir():
            continue

        # Extract student ID from folder name (e.g., Participant_38951_assignsubmission_file)
        match = re.search(r'Participant_(\d+)_', folder.name)
        if match:
            student_id = match.group(1)
            folders[student_id] = folder

    print(f"Found {len(folders)} student folders\n")
    return folders


def extract_actual_grades(student_folders, evaluator_grades):
    """Extract actual grades from PDFs for each student"""

    print("Extracting actual grades from PDFs...")

    results = []
    found_count = 0
    not_found_count = 0
    parse_failed_count = 0

    for i, (student_id, data) in enumerate(evaluator_grades.items(), 1):
        result = {
            'student_id': student_id,
            'evaluator_grade': data['final_grade'],
            'base_pct': data['base_pct'],
            'bonus': data['bonus'],
            'criteria_count': data['criteria_count'],
            'rank': data['rank'],
            'actual_grade': None,
            'status': 'Not Found'
        }

        # Find student folder
        if student_id not in student_folders:
            print(f"[{i}/{len(evaluator_grades)}] Student {student_id}: Folder not found")
            not_found_count += 1
            results.append(result)
            continue

        folder = student_folders[student_id]

        # Find PDF
        pdf_name = f"Detailed_Grade_Breakdown_{student_id}.pdf"
        pdf_path = folder / pdf_name

        if not pdf_path.exists():
            print(f"[{i}/{len(evaluator_grades)}] Student {student_id}: PDF not found")
            result['status'] = 'No PDF'
            not_found_count += 1
            results.append(result)
            continue

        # Extract grade
        actual_grade = extract_grade_from_pdf(pdf_path)

        if actual_grade is None:
            print(f"[{i}/{len(evaluator_grades)}] Student {student_id}: Could not parse grade")
            result['status'] = 'Parse Failed'
            parse_failed_count += 1
            results.append(result)
            continue

        result['actual_grade'] = actual_grade
        result['status'] = 'Success'
        difference = result['evaluator_grade'] - actual_grade

        print(f"[{i}/{len(evaluator_grades)}] Student {student_id}: "
              f"Actual={actual_grade:.1f}, Evaluator={result['evaluator_grade']:.1f}, "
              f"Diff={difference:+.1f}")

        found_count += 1
        results.append(result)

    print(f"\nSuccessfully extracted: {found_count}/{len(evaluator_grades)}")
    print(f"Not found: {not_found_count}, Parse failed: {parse_failed_count}\n")

    return results


def calculate_statistics(results):
    """Calculate comparison statistics"""

    # Filter results with valid actual grades
    valid_results = [r for r in results if r['actual_grade'] is not None]

    if not valid_results:
        return None

    evaluator_grades = [r['evaluator_grade'] for r in valid_results]
    actual_grades = [r['actual_grade'] for r in valid_results]

    differences = [e - a for e, a in zip(evaluator_grades, actual_grades)]
    abs_differences = [abs(d) for d in differences]
    percentage_errors = [(abs(e - a) / a) * 100 for e, a in zip(evaluator_grades, actual_grades) if a != 0]

    # Calculate correlation
    import numpy as np
    correlation = np.corrcoef(evaluator_grades, actual_grades)[0, 1] if len(valid_results) > 1 else 0

    # Calculate RMSE
    rmse = np.sqrt(np.mean([d**2 for d in differences]))

    # Count students within ranges
    within_5 = sum(1 for d in abs_differences if d <= 5)
    within_10 = sum(1 for d in abs_differences if d <= 10)

    # Count over/under grading
    over_graded = sum(1 for d in differences if d > 0)
    under_graded = sum(1 for d in differences if d < 0)

    stats = {
        'count': len(valid_results),
        'mean_evaluator': np.mean(evaluator_grades),
        'mean_actual': np.mean(actual_grades),
        'mean_difference': np.mean(differences),
        'mean_abs_difference': np.mean(abs_differences),
        'mean_percentage_error': np.mean(percentage_errors) if percentage_errors else 0,
        'correlation': correlation,
        'rmse': rmse,
        'within_5': within_5,
        'within_10': within_10,
        'over_graded': over_graded,
        'under_graded': under_graded,
    }

    return stats


def generate_excel_report(results, stats, output_path):
    """Generate comprehensive Excel comparison report"""

    print("Generating Excel report...")

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    good_fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
    warning_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    bad_fill = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ========================================================================
    # SHEET 1: COMPARISON
    # ========================================================================
    ws1 = wb.active
    ws1.title = "Comparison"

    headers = ["Rank", "Student ID", "Evaluator Grade", "Actual Grade", "Difference",
               "Abs Diff", "% Error", "Criteria", "Status"]

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Add data
    valid_results = [r for r in results if r['actual_grade'] is not None]

    for i, result in enumerate(valid_results, 2):
        diff = result['evaluator_grade'] - result['actual_grade']
        abs_diff = abs(diff)
        pct_error = (abs_diff / result['actual_grade']) * 100 if result['actual_grade'] != 0 else 0

        ws1.cell(row=i, column=1, value=result['rank'])
        ws1.cell(row=i, column=2, value=result['student_id'])
        ws1.cell(row=i, column=3, value=round(result['evaluator_grade'], 1))
        ws1.cell(row=i, column=4, value=round(result['actual_grade'], 1))
        ws1.cell(row=i, column=5, value=round(diff, 1))
        ws1.cell(row=i, column=6, value=round(abs_diff, 1))
        ws1.cell(row=i, column=7, value=f"{pct_error:.1f}%")
        ws1.cell(row=i, column=8, value=result['criteria_count'])
        ws1.cell(row=i, column=9, value=result['status'])

        # Color code by difference
        for col in range(1, 10):
            cell = ws1.cell(row=i, column=col)
            cell.border = thin_border

            if abs_diff <= 5:
                cell.fill = good_fill
            elif abs_diff <= 10:
                cell.fill = warning_fill
            else:
                cell.fill = bad_fill

    # Auto-width columns
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws1.column_dimensions[column].width = max_length + 2

    # ========================================================================
    # SHEET 2: STATISTICS
    # ========================================================================
    ws2 = wb.create_sheet("Statistics")

    if stats:
        ws2.cell(row=1, column=1, value="Metric").font = header_font
        ws2.cell(row=1, column=2, value="Value").font = header_font

        metrics = [
            ("Total Students Compared", stats['count']),
            ("", ""),
            ("Mean Evaluator Grade", f"{stats['mean_evaluator']:.1f}"),
            ("Mean Actual Grade", f"{stats['mean_actual']:.1f}"),
            ("", ""),
            ("Mean Difference", f"{stats['mean_difference']:+.1f}"),
            ("Mean Absolute Error", f"{stats['mean_abs_difference']:.1f}"),
            ("Mean Percentage Error", f"{stats['mean_percentage_error']:.1f}%"),
            ("", ""),
            ("Correlation Coefficient", f"{stats['correlation']:.3f}"),
            ("RMSE", f"{stats['rmse']:.1f}"),
            ("", ""),
            ("Students within ±5 points", f"{stats['within_5']}/{stats['count']} ({stats['within_5']/stats['count']*100:.1f}%)"),
            ("Students within ±10 points", f"{stats['within_10']}/{stats['count']} ({stats['within_10']/stats['count']*100:.1f}%)"),
            ("", ""),
            ("Over-graded (evaluator > actual)", stats['over_graded']),
            ("Under-graded (evaluator < actual)", stats['under_graded']),
        ]

        for i, (metric, value) in enumerate(metrics, 2):
            ws2.cell(row=i, column=1, value=metric)
            ws2.cell(row=i, column=2, value=value)

        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 20

    # ========================================================================
    # SHEET 3: DISCREPANCIES
    # ========================================================================
    ws3 = wb.create_sheet("Discrepancies")

    # Filter for large discrepancies (>15 points)
    large_discrepancies = [r for r in valid_results
                          if abs(r['evaluator_grade'] - r['actual_grade']) > 15]

    headers = ["Student ID", "Evaluator", "Actual", "Difference", "Type"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, result in enumerate(sorted(large_discrepancies,
                                     key=lambda x: abs(x['evaluator_grade'] - x['actual_grade']),
                                     reverse=True), 2):
        diff = result['evaluator_grade'] - result['actual_grade']
        discrepancy_type = "Over-graded" if diff > 0 else "Under-graded"

        ws3.cell(row=i, column=1, value=result['student_id'])
        ws3.cell(row=i, column=2, value=round(result['evaluator_grade'], 1))
        ws3.cell(row=i, column=3, value=round(result['actual_grade'], 1))
        ws3.cell(row=i, column=4, value=round(diff, 1))
        ws3.cell(row=i, column=5, value=discrepancy_type)

        for col in range(1, 6):
            ws3.cell(row=i, column=col).border = thin_border

    for col in ws3.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws3.column_dimensions[column].width = max_length + 2

    # Save workbook
    wb.save(output_path)
    print(f"Saved: {output_path}")


def print_summary(stats):
    """Print summary statistics to console"""

    if not stats:
        print("No valid comparisons to summarize.")
        return

    print("=" * 80)
    print("SUMMARY STATISTICS")
    print("=" * 80)
    print()
    print(f"Mean Evaluator Grade:   {stats['mean_evaluator']:.1f}")
    print(f"Mean Actual Grade:      {stats['mean_actual']:.1f}")
    print(f"Mean Difference:        {stats['mean_difference']:+.1f} ", end="")
    if stats['mean_difference'] > 0:
        print("(evaluator grades HIGHER on average)")
    elif stats['mean_difference'] < 0:
        print("(evaluator grades LOWER on average)")
    else:
        print("(perfectly calibrated)")

    print(f"Mean Absolute Error:    {stats['mean_abs_difference']:.1f}")
    print(f"Correlation:            {stats['correlation']:.3f} ", end="")

    if stats['correlation'] >= 0.9:
        print("(excellent correlation)")
    elif stats['correlation'] >= 0.7:
        print("(strong correlation)")
    elif stats['correlation'] >= 0.5:
        print("(moderate correlation)")
    else:
        print("(weak correlation)")

    print(f"RMSE:                   {stats['rmse']:.1f}")
    print()
    print(f"Students within ±5 points:   {stats['within_5']}/{stats['count']} ({stats['within_5']/stats['count']*100:.1f}%)")
    print(f"Students within ±10 points:  {stats['within_10']}/{stats['count']} ({stats['within_10']/stats['count']*100:.1f}%)")
    print()
    print("=" * 80)
    print()


def main():
    """Main execution"""

    if len(sys.argv) < 2:
        print("Usage: python compare_grades.py <worksubmissions_path>")
        print("Example: python compare_grades.py E:\\Projects\\student-project-evaluator\\tests\\WorkSubmissions01")
        sys.exit(1)

    worksubmissions_path = sys.argv[1]

    print("=" * 80)
    print("GRADE COMPARISON TOOL")
    print("=" * 80)
    print()

    # Paths
    grades_path = "outputs/grades.xlsx"
    output_path = "outputs/grade_comparison.xlsx"

    # Load evaluator grades
    if not os.path.exists(grades_path):
        print(f"Error: Could not find {grades_path}")
        print("Please run student-project-evaluator first.")
        sys.exit(1)

    evaluator_grades = load_evaluator_grades(grades_path)

    # Find student folders
    student_folders = find_student_folders(worksubmissions_path)

    # Extract actual grades
    results = extract_actual_grades(student_folders, evaluator_grades)

    # Calculate statistics
    stats = calculate_statistics(results)

    # Print summary
    print_summary(stats)

    # Generate Excel report
    generate_excel_report(results, stats, output_path)

    print()


if __name__ == "__main__":
    main()
