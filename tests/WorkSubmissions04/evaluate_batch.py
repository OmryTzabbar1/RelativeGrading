#!/usr/bin/env python3
"""
Batch evaluation script for WorkSubmissions04
Evaluates 20 student projects by extracting criteria from markdown files
"""

import os
import json
import re
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

# Student IDs (extracted from folder names)
STUDENT_FOLDERS = [
    "Participant_87681_assignsubmission_file",
    "Participant_87682_assignsubmission_file",
    "Participant_87683_assignsubmission_file",
    "Participant_87684_assignsubmission_file",
    "Participant_87685_assignsubmission_file",
    "Participant_87687_assignsubmission_file",
    "Participant_87690_assignsubmission_file",
    "Participant_87691_assignsubmission_file",
    "Participant_87692_assignsubmission_file",
    "Participant_87694_assignsubmission_file",
    "Participant_87696_assignsubmission_file",
    "Participant_87698_assignsubmission_file",
    "Participant_87703_assignsubmission_file",
    "Participant_87706_assignsubmission_file",
    "Participant_87707_assignsubmission_file",
    "Participant_87714_assignsubmission_file",
    "Participant_87715_assignsubmission_file",
    "Participant_87716_assignsubmission_file",
    "Participant_87719_assignsubmission_file",
    "Participant_87721_assignsubmission_file",
]

# Category keywords (from CATEGORIES.md)
CATEGORY_KEYWORDS = {
    "Documentation": [
        "readme", "api docs", "api documentation", "user guide", "usage guide",
        "changelog", "change log", "contributing", "contribution guide", "license",
        "installation guide", "setup guide", "getting started", "faq",
        "frequently asked questions", "wiki", "manual", "reference docs"
    ],
    "Planning": [
        "prd", "product requirements", "architecture", "architectural design",
        "design document", "design doc", "technical spec", "technical specification",
        "roadmap", "project roadmap", "milestones", "requirements", "requirements document",
        "functional requirements", "non-functional requirements", "system design",
        "database schema", "db schema", "wireframes", "mockups", "user stories",
        "use cases", "use case", "problem statement", "problem", "solution overview",
        "proposed solution", "project goals", "objectives", "success metrics", "kpis",
        "evaluation criteria", "assumptions", "assumptions documentation", "constraints",
        "limitations", "scope", "design decisions", "design rationale", "trade-offs",
        "alternatives considered", "technology selection", "technology justification",
        "user personas", "target users", "user flow", "user journey", "timeline",
        "schedule", "project plan"
    ],
    "Testing": [
        "unit test", "unit tests", "unit testing", "integration test", "integration tests",
        "e2e test", "e2e tests", "end-to-end", "test coverage", "code coverage",
        "pytest", "jest", "mocha", "junit", "testing framework", "test suite",
        "automated tests", "automated testing", "manual testing", "regression tests",
        "smoke tests", "load testing", "performance testing", "test cases", "tdd",
        "test-driven"
    ],
    "DevOps": [
        "ci/cd", "ci cd", "continuous integration", "continuous deployment",
        "github actions", "gitlab ci", "jenkins", "docker", "dockerfile",
        "containerization", "kubernetes", "k8s", "deployment", "deploy script",
        "aws", "azure", "gcp", "cloud", "terraform", "infrastructure as code",
        "monitoring", "logging", "nginx", "apache", "load balancer", "ssl", "https",
        "environment variables", "env config", "build pipeline", "release automation"
    ],
    "Research": [
        "research", "research findings", "analysis", "data analysis", "jupyter notebook",
        "notebook", "data exploration", "experiment", "experimentation", "hypothesis",
        "findings", "results", "literature review", "benchmarking", "benchmark",
        "comparison study", "survey", "user research", "insights", "statistical analysis",
        "machine learning experiments", "model evaluation"
    ],
    "Visuals": [
        "screenshot", "screenshots", "diagram", "diagrams", "flowchart", "flow chart",
        "chart", "charts", "graph", "graphs", "demo video", "video demo", "gif",
        "animated gif", "mockup", "wireframe", "ui preview", "before/after",
        "architecture diagram", "sequence diagram", "erd", "entity relationship diagram",
        "class diagram", "infographic"
    ],
    "CodeQuality": [
        "linting", "linter", "eslint", "pylint", "ruff", "flake8",
        "prettier", "code formatting", "black", "autopep8", "yapf",
        "type checking", "typescript", "mypy", "pyright", "pyre",
        "code review", "peer review", "pre-commit", "git hooks", "husky", "lint-staged",
        "refactoring", "clean code", "solid principles", "design patterns",
        "code style", "style guide", "coding standards", "pep 8", "pep8", "airbnb style", "google style",
        "documentation strings", "docstrings", "comments", "code comments",
        "static analysis", "sonarqube", "sonarcloud",
        "code complexity", "technical debt", "quality gates", "quality checks",
        "project setup", "setup.py", "pyproject.toml", "package.json",
        "dependency management", "requirements", "environment management", "venv", "virtual environment"
    ],
    "Business": [
        "cost analysis", "cost breakdown", "budget", "roi", "return on investment",
        "market research", "market analysis", "user personas", "customer personas",
        "business case", "business model", "pricing", "pricing strategy",
        "monetization", "competitive analysis", "competitor comparison",
        "swot analysis", "value proposition", "target audience",
        "stakeholder analysis", "budgeting", "revenue model", "go-to-market",
        "risk analysis", "risk assessment", "risks", "risk mitigation"
    ]
}

# Filename patterns that auto-credit criteria (from EXTRACTION.md)
FILENAME_CRITERIA = {
    r"^(PRD|prd|ProductRequirements)\.md$": "PRD Document",
    r"^(TESTING|testing|TEST)\.md$": "Testing Documentation",
    r"^(CONTRIBUTING|contributing)\.md$": "Contributing Guide",
    r"^(QUICKSTART|QuickStart|quick-start)\.md$": "Quick Start Guide",
    r"^(ARCHITECTURE|architecture|DESIGN)\.md$": "Architecture Documentation",
    r"^(CHANGELOG|changelog|CHANGES)\.md$": "Changelog",
    r"^(API|api|API_DOCS)\.md$": "API Documentation",
    r"^(ROADMAP|roadmap)\.md$": "Roadmap",
    r"^(DEPLOYMENT|deployment|DEPLOY)\.md$": "Deployment Guide",
    r"^(TROUBLESHOOTING|troubleshooting|FAQ)\.md$": "Troubleshooting Guide",
    r"^(README|readme)\.md$": "README",
}

# Implementation indicators (from EXTRACTION.md)
IMPLEMENTATION_INDICATORS = [
    r"\bwe built\b", r"\bwe created\b", r"\bwe developed\b", r"\bwe implemented\b",
    r"\bproject includes\b", r"\bthis project includes\b", r"\bincludes\b",
    r"\bimplemented\b", r"\bcreated\b", r"\bdeveloped\b", r"\bbuilt\b",
    r"\badded\b", r"\bintegrated\b", r"\bincorporated\b",
    r"\bsupports\b", r"\bprovides\b", r"\boffers\b", r"\benables\b",
    r"\bhas\b", r"\bcontains\b", r"\bcomes with\b",
    r"\bcomplete\b", r"\bcomprehensive\b", r"\bfull\b",
    r"\bavailable\b", r"\bfeaturing\b",
    r"\bwith support for\b", r"\bcapable of\b",
    r"\btotal tests\b", r"\btest cases\b", r"\btests\b",
    r"\bcoverage\b", r"\bcode coverage\b", r"\btest coverage\b",
]

# Negative indicators (from EXTRACTION.md)
NEGATIVE_INDICATORS = [
    r"\bTODO\b", r"\bFIXME\b", r"\bwill add\b", r"\bplanning to\b",
    r"\bplanned for\b", r"\bfuture work\b", r"\bnext steps\b",
    r"\bnot yet implemented\b", r"\bnot complete\b", r"\bmissing\b",
    r"\bout of scope\b", r"\bskipped\b", r"\bomitted\b",
    r"\bno .* currently\b", r"\bnot working\b",
    r"\bmight be added\b", r"\bconsidering\b", r"\bif time permits\b",
    r"\bwould be nice\b", r"\bideally\b", r"\boptional\b",
]


def normalize_criterion_name(name):
    """Normalize criterion names for consistency"""
    # Title case, but preserve acronyms
    words = name.split()
    normalized = []
    for word in words:
        if word.upper() == word and len(word) > 1:  # Acronym
            normalized.append(word)
        else:
            normalized.append(word.capitalize())
    return " ".join(normalized)


def categorize_criterion(criterion_name):
    """Categorize a criterion based on keywords"""
    name_lower = criterion_name.lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        for keyword in keywords:
            if keyword in name_lower:
                return category
    return "Uncategorized"


def check_filename_criteria(filename):
    """Check if filename matches any auto-credit patterns"""
    for pattern, criterion in FILENAME_CRITERIA.items():
        if re.match(pattern, filename, re.IGNORECASE):
            return criterion
    return None


def has_negative_context(text, position):
    """Check if text around position has negative indicators"""
    # Check 200 chars before and after
    start = max(0, position - 200)
    end = min(len(text), position + 200)
    context = text[start:end].lower()

    for pattern in NEGATIVE_INDICATORS:
        if re.search(pattern, context, re.IGNORECASE):
            return True
    return False


def extract_section_headers(content):
    """Extract all markdown section headers"""
    headers = []
    for match in re.finditer(r'^#{1,6}\s+(.+)$', content, re.MULTILINE):
        headers.append(match.group(1).strip())
    return headers


def extract_criteria_from_content(content, filepath):
    """Extract criteria from markdown content"""
    criteria = set()
    content_lower = content.lower()

    # Extract section headers
    headers = extract_section_headers(content)

    # Map headers to criteria
    header_mappings = {
        r"testing|tests": "Testing Documentation",
        r"unit tests?": "Unit Tests",
        r"installation|setup|getting started": "Installation Instructions",
        r"usage|how to use": "Usage Guide",
        r"features?|functionality": None,  # Extract features from content
        r"screenshots?|demo|examples": "Screenshots",
        r"architecture|design": "Architecture Documentation",
        r"ci/cd|deployment|devops": "DevOps Documentation",
        r"api|endpoints|api reference": "API Documentation",
        r"contributing|development": "Contributing Guide",
        r"problem statement|problem": "Problem Statement",
        r"solution|proposed solution": "Solution Overview",
        r"requirements?|functional requirements": "Requirements Documentation",
        r"use cases?": "Use Case Documentation",
        r"cost analysis|budget|costs": "Cost Analysis",
        r"assumptions?|constraints": "Assumptions Documentation",
        r"success metrics|kpis|metrics": "Success Metrics",
        r"risk analysis|risks": "Risk Analysis",
        r"roadmap|future work|next steps": "Roadmap",
        r"user stories|user personas": "User Research",
        r"design decisions|trade-offs": "Design Rationale",
    }

    for header in headers:
        header_lower = header.lower()
        for pattern, criterion in header_mappings.items():
            if re.search(pattern, header_lower):
                if criterion:
                    criteria.add(criterion)
                # For features section, parse the content below
                break

    # Direct Quality Standards scanning (check for mentions throughout content)
    quality_patterns = [
        (r"eslint|\.eslintrc", "ESLint Configuration"),
        (r"pylint|\.pylintrc", "Pylint Configuration"),
        (r"\bruff\b", "Ruff Linting"),
        (r"prettier|\.prettierrc", "Prettier Formatting"),
        (r"black formatter|black formatting|formatted with black|\bblack\b.*format", "Black Formatting"),
        (r"pre-commit|\.pre-commit-config", "Pre-commit Hooks"),
        (r"git hooks|husky|lint-staged", "Git Hooks"),
        (r"typescript.*strict|strict.*typescript", "TypeScript Type Checking"),
        (r"mypy|type hints|type annotations|pyright|pyre", "Mypy Type Checking"),
        (r"pep ?8|pep8", "PEP8 Compliance"),
        (r"code style guide|style guide|coding standards|airbnb style|google style", "Code Style Guide"),
        (r"code review|peer review|pull request review", "Code Review Process"),
        (r"quality gates|quality checks|automated quality", "Quality Gates"),
        (r"flake8", "Flake8 Linting"),
        (r"sonarqube|sonarcloud", "Static Analysis Tools"),
    ]

    for pattern, criterion_name in quality_patterns:
        matches = list(re.finditer(pattern, content_lower))
        for match in matches:
            # Check if not in negative context
            if not has_negative_context(content_lower, match.start()):
                criteria.add(criterion_name)
                break  # Only add once per pattern

    # Look for implementation indicators
    for indicator_pattern in IMPLEMENTATION_INDICATORS:
        for match in re.finditer(indicator_pattern, content_lower):
            # Check for negative context
            if not has_negative_context(content_lower, match.start()):
                # Try to extract what follows (e.g., "includes unit tests")
                pos = match.end()
                snippet = content_lower[pos:pos+100]

                # Common feature patterns
                if re.search(r"unit tests?|testing", snippet):
                    criteria.add("Unit Tests")
                if re.search(r"integration tests?", snippet):
                    criteria.add("Integration Tests")
                if re.search(r"e2e|end-to-end", snippet):
                    criteria.add("E2E Tests")
                if re.search(r"coverage|code coverage", snippet):
                    criteria.add("Test Coverage Metrics")
                if re.search(r"ci/cd|continuous integration", snippet):
                    criteria.add("CI/CD Pipeline")
                if re.search(r"docker|containeriz", snippet):
                    criteria.add("Docker Containerization")
                if re.search(r"frontend|front-end|ui", snippet):
                    criteria.add("Frontend Implementation")
                if re.search(r"backend|back-end|api", snippet):
                    criteria.add("Backend Implementation")
                if re.search(r"database|db", snippet):
                    criteria.add("Database Integration")
                if re.search(r"authentication|auth", snippet):
                    criteria.add("Authentication System")
                if re.search(r"real-time|realtime", snippet):
                    criteria.add("Real-time Features")

                # Quality Standards extraction
                if re.search(r"eslint|\.eslintrc", snippet):
                    criteria.add("ESLint Configuration")
                if re.search(r"pylint|\.pylintrc", snippet):
                    criteria.add("Pylint Configuration")
                if re.search(r"ruff", snippet):
                    criteria.add("Ruff Linting")
                if re.search(r"prettier|\.prettierrc", snippet):
                    criteria.add("Prettier Formatting")
                if re.search(r"black formatter|black formatting|formatted with black", snippet):
                    criteria.add("Black Formatting")
                if re.search(r"pre-commit|\.pre-commit-config", snippet):
                    criteria.add("Pre-commit Hooks")
                if re.search(r"git hooks|husky", snippet):
                    criteria.add("Git Hooks")
                if re.search(r"typescript.*strict|strict.*typescript|type checking", snippet):
                    criteria.add("TypeScript Type Checking")
                if re.search(r"mypy|type hints|type annotations", snippet):
                    criteria.add("Mypy Type Checking")
                if re.search(r"pep ?8|pep8", snippet):
                    criteria.add("PEP8 Compliance")
                if re.search(r"code style guide|style guide|coding standards", snippet):
                    criteria.add("Code Style Guide")
                if re.search(r"code review|peer review", snippet):
                    criteria.add("Code Review Process")
                if re.search(r"quality gates|quality checks", snippet):
                    criteria.add("Quality Gates")

    # Extract coverage percentages
    coverage_matches = re.findall(r'(\d+)%\s*(?:code\s*)?coverage', content_lower)
    if coverage_matches:
        criteria.add("Test Coverage Metrics")

    # Extract test counts
    test_count_patterns = [
        r'(\d+)\+?\s*tests',
        r'total tests:\s*~?(\d+)',
        r'~(\d+)\s*test cases',
    ]
    for pattern in test_count_patterns:
        if re.search(pattern, content_lower):
            criteria.add("Unit Tests")
            break

    return criteria


def extract_student_criteria(student_folder):
    """Extract all criteria for a student"""
    student_path = BASE_DIR / student_folder
    criteria = set()
    files_processed = []

    # Find all .md files
    for md_file in student_path.rglob("*.md"):
        files_processed.append(str(md_file.relative_to(student_path)))

        # Check filename
        filename = md_file.name
        filename_criterion = check_filename_criteria(filename)
        if filename_criterion:
            criteria.add(filename_criterion)

        # Read content
        try:
            with open(md_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()

            # Extract criteria from content
            content_criteria = extract_criteria_from_content(content, md_file)
            criteria.update(content_criteria)
        except Exception as e:
            print(f"Error reading {md_file}: {e}")

    return criteria, files_processed


def build_criteria_graph(student_data):
    """Build criteria graph with weights and categories"""
    criteria_graph = defaultdict(lambda: {
        "display_name": "",
        "students": [],
        "count": 0,
        "weight": 0.0,
        "category": "Uncategorized"
    })

    total_students = len(student_data)

    # Aggregate criteria across students
    for student_id, data in student_data.items():
        for criterion in data["criteria"]:
            normalized = normalize_criterion_name(criterion)
            criteria_graph[normalized]["display_name"] = normalized
            criteria_graph[normalized]["students"].append(student_id)
            criteria_graph[normalized]["count"] += 1

    # Calculate weights and categorize
    for criterion, data in criteria_graph.items():
        data["weight"] = data["count"] / total_students
        data["category"] = categorize_criterion(criterion)

    return dict(criteria_graph)


def calculate_scores(student_data, criteria_graph):
    """Calculate scores for each student"""
    scores = {}

    for student_id, data in student_data.items():
        raw_score = 0.0
        rarity_bonus = 0.0

        for criterion in data["criteria"]:
            normalized = normalize_criterion_name(criterion)
            weight = criteria_graph[normalized]["weight"]
            raw_score += weight

            # Rarity bonus: ≤15% prevalence (≤3 students out of 20)
            if criteria_graph[normalized]["count"] <= 3:
                rarity_bonus += 1.0

        total_score = raw_score + rarity_bonus

        scores[student_id] = {
            "raw_score": raw_score,
            "rarity_bonus": rarity_bonus,
            "total_score": total_score,
            "criteria_count": len(data["criteria"])
        }

    # Calculate max possible score
    max_possible = sum(c["weight"] for c in criteria_graph.values())

    # Calculate percentages
    for student_id in scores:
        scores[student_id]["max_possible"] = max_possible
        scores[student_id]["percentage"] = (scores[student_id]["raw_score"] / max_possible) * 100 if max_possible > 0 else 0

    return scores


def assign_relative_grades(scores):
    """Assign relative grades (best student = 100)"""
    # Find best percentage
    best_percentage = max(s["percentage"] for s in scores.values())

    # Calculate relative grades
    for student_id in scores:
        student_pct = scores[student_id]["percentage"]
        scores[student_id]["grade"] = (student_pct / best_percentage) * 100 if best_percentage > 0 else 0

    # Sort by grade
    sorted_students = sorted(scores.items(), key=lambda x: x[1]["grade"], reverse=True)

    # Assign ranks
    for rank, (student_id, _) in enumerate(sorted_students, 1):
        scores[student_id]["rank"] = rank

    return scores


def create_excel_report(student_data, criteria_graph, scores, output_path):
    """Create comprehensive Excel report"""
    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ["Student Project Evaluation Report"],
        ["Evaluation Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Students", len(student_data)],
        ["Total Criteria", len(criteria_graph)],
        [""],
        ["Top 3 Students"],
        ["Rank", "Student ID", "Grade", "Score", "Criteria Count"],
    ]

    sorted_students = sorted(scores.items(), key=lambda x: x[1]["rank"])
    for student_id, data in sorted_students[:3]:
        summary_data.append([
            data["rank"],
            student_id,
            f"{data['grade']:.2f}",
            f"{data['total_score']:.2f}",
            data["criteria_count"]
        ])

    for row in summary_data:
        ws_summary.append(row)

    # Style summary
    ws_summary["A1"].font = Font(size=16, bold=True)
    ws_summary.merge_cells("A1:E1")

    # Sheet 2: Grades
    ws_grades = wb.create_sheet("Grades")
    ws_grades.append(["Student ID", "Raw Score", "Rarity Bonus", "Total Score", "Max Possible", "Percentage", "Grade", "Rank", "Criteria Count"])

    for i, cell in enumerate(ws_grades[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for student_id, data in sorted(scores.items(), key=lambda x: x[1]["rank"]):
        ws_grades.append([
            student_id,
            round(data["raw_score"], 2),
            round(data["rarity_bonus"], 2),
            round(data["total_score"], 2),
            round(data["max_possible"], 2),
            round(data["percentage"], 2),
            round(data["grade"], 2),
            data["rank"],
            data["criteria_count"]
        ])

    # Sheet 3: Criteria Distribution
    ws_criteria = wb.create_sheet("Criteria Distribution")
    ws_criteria.append(["Criterion", "Count", "Weight", "Category", "Students"])

    for i, cell in enumerate(ws_criteria[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for criterion, data in sorted(criteria_graph.items(), key=lambda x: x[1]["count"], reverse=True):
        ws_criteria.append([
            criterion,
            data["count"],
            round(data["weight"], 3),
            data["category"],
            ", ".join(data["students"])
        ])

    # Sheet 4: Category Breakdown
    ws_categories = wb.create_sheet("Category Breakdown")
    ws_categories.append(["Category", "Total Criteria", "Average Count per Student"])

    for i, cell in enumerate(ws_categories[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    category_stats = defaultdict(lambda: {"total": 0, "sum": 0})
    for criterion, data in criteria_graph.items():
        cat = data["category"]
        category_stats[cat]["total"] += 1
        category_stats[cat]["sum"] += data["count"]

    for category, stats in sorted(category_stats.items()):
        avg_count = stats["sum"] / len(student_data)
        ws_categories.append([category, stats["total"], round(avg_count, 2)])

    # Sheet 5: Student Details
    ws_details = wb.create_sheet("Student Details")
    ws_details.append(["Student ID", "Rank", "Grade", "Criteria"])

    for i, cell in enumerate(ws_details[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for student_id, data in sorted(student_data.items(), key=lambda x: scores[x[0]]["rank"]):
        criteria_list = ", ".join(sorted(normalize_criterion_name(c) for c in data["criteria"]))
        ws_details.append([
            student_id,
            scores[student_id]["rank"],
            round(scores[student_id]["grade"], 2),
            criteria_list
        ])

    # Sheet 6: Flagged Criteria (rare criteria)
    ws_flagged = wb.create_sheet("Flagged Criteria")
    ws_flagged.append(["Criterion", "Count", "Weight", "Category", "Rarity Bonus", "Students"])

    for i, cell in enumerate(ws_flagged[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for criterion, data in sorted(criteria_graph.items(), key=lambda x: x[1]["count"]):
        if data["count"] <= 3:  # Rare criteria
            ws_flagged.append([
                criterion,
                data["count"],
                round(data["weight"], 3),
                data["category"],
                "Yes (+1)",
                ", ".join(data["students"])
            ])

    # Auto-size columns
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)


def main():
    print("=" * 80)
    print("STUDENT PROJECT EVALUATOR - WorkSubmissions04")
    print("=" * 80)
    print()

    # Extract criteria for all students
    print("Phase 1: Extracting criteria from student projects...")
    student_data = {}

    for folder in STUDENT_FOLDERS:
        student_id = folder.split("_")[1]  # Extract ID (e.g., "87681")
        print(f"  Processing {student_id}...", end=" ")

        criteria, files = extract_student_criteria(folder)
        student_data[student_id] = {
            "criteria": criteria,
            "files_processed": files
        }
        print(f"{len(criteria)} criteria from {len(files)} files")

    print()

    # Build criteria graph
    print("Phase 2: Building criteria graph...")
    criteria_graph = build_criteria_graph(student_data)
    print(f"  Total unique criteria: {len(criteria_graph)}")
    print()

    # Calculate scores
    print("Phase 3: Calculating scores...")
    scores = calculate_scores(student_data, criteria_graph)
    scores = assign_relative_grades(scores)
    print(f"  Scores calculated for {len(scores)} students")
    print()

    # Generate outputs
    print("Phase 4: Generating output files...")

    # Save criteria graph JSON
    criteria_json_path = OUTPUT_DIR / "criteria_graph_final.json"
    with open(criteria_json_path, 'w') as f:
        json.dump({
            "metadata": {
                "total_students": len(student_data),
                "evaluation_date": datetime.now().isoformat(),
                "total_criteria": len(criteria_graph)
            },
            "criteria": criteria_graph
        }, f, indent=2)
    print(f"  [OK] Saved {criteria_json_path}")

    # Save grades Excel
    grades_excel_path = OUTPUT_DIR / "grades.xlsx"
    wb_grades = openpyxl.Workbook()
    ws = wb_grades.active
    ws.title = "Grades"
    ws.append(["Student ID", "Raw Score", "Max Possible", "Percentage", "Grade", "Rank", "Criteria Count"])

    for student_id, data in sorted(scores.items(), key=lambda x: x[1]["rank"]):
        ws.append([
            student_id,
            round(data["raw_score"], 2),
            round(data["max_possible"], 2),
            round(data["percentage"], 2),
            round(data["grade"], 2),
            data["rank"],
            data["criteria_count"]
        ])

    wb_grades.save(grades_excel_path)
    print(f"  [OK] Saved {grades_excel_path}")

    # Save comprehensive report
    report_excel_path = OUTPUT_DIR / "Student_Evaluation_Report.xlsx"
    create_excel_report(student_data, criteria_graph, scores, report_excel_path)
    print(f"  [OK] Saved {report_excel_path}")
    print()

    # Print summary
    print("=" * 80)
    print("EVALUATION SUMMARY")
    print("=" * 80)
    print(f"Total Students Evaluated: {len(student_data)}")
    print(f"Total Criteria Discovered: {len(criteria_graph)}")
    print()

    print("Top 3 Students:")
    for student_id, data in sorted(scores.items(), key=lambda x: x[1]["rank"])[:3]:
        print(f"  {data['rank']}. Student {student_id}: Grade {data['grade']:.2f} ({data['criteria_count']} criteria)")
    print()

    # Grade distribution
    print("Grade Distribution:")
    ranges = [(90, 100), (80, 90), (70, 80), (60, 70), (0, 60)]
    for low, high in ranges:
        count = sum(1 for s in scores.values() if low <= s["grade"] < high)
        print(f"  {low}-{high}: {count} students")
    print()

    # Planning & Analysis criteria
    print("Planning & Analysis Criteria Found:")
    planning_criteria = [(c, d) for c, d in criteria_graph.items() if d["category"] == "Planning"]
    planning_criteria.sort(key=lambda x: x[1]["count"], reverse=True)
    for criterion, data in planning_criteria[:10]:
        print(f"  - {criterion}: {data['count']} students")

    print()

    # Code Quality criteria
    print("Code Quality Criteria Found:")
    quality_criteria = [(c, d) for c, d in criteria_graph.items() if d["category"] == "CodeQuality"]
    quality_criteria.sort(key=lambda x: x[1]["count"], reverse=True)
    if quality_criteria:
        for criterion, data in quality_criteria:
            print(f"  - {criterion}: {data['count']} students")
    else:
        print("  (none found)")
    print()

    # Category breakdown
    print("Criteria by Category:")
    category_counts = defaultdict(int)
    for criterion, data in criteria_graph.items():
        category_counts[data["category"]] += 1
    for category in sorted(category_counts.keys()):
        print(f"  {category}: {category_counts[category]} criteria")

    print()
    print("=" * 80)
    print("Evaluation complete!")
    print("=" * 80)


if __name__ == "__main__":
    main()
