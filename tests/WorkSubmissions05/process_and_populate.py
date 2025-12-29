#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Process WorkSubmissions05 PDFs and create individual submission_info.xlsx files
Extracts grade information from first 2 pages of each PDF and populates Excel in each folder
"""

import sys
import codecs
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Set UTF-8 encoding for stdout on Windows
if sys.platform == 'win32':
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')


def extract_participant_id(folder_name):
    """Extract participant ID from folder name."""
    match = re.search(r'Participant_(\d+)', folder_name)
    return match.group(1) if match else folder_name


def extract_info_from_pdf_text(pdf_text):
    """
    Extract the required pieces of information from PDF text.
    Looks at first ~2 pages (passed as pdf_text parameter).
    """
    info = {
        'group_code': '',
        'student1': '',
        'student2': '',
        'github': '',
        'grade': ''
    }

    # Patterns for extraction
    patterns = {
        'group_code': [
            r'1\.\s*(?:Group [Cc]ode|Codegroupe|קוד קבוצה):\s*([^\n]+)',
            r'Group [Cc]ode:\s*([^\n]+)',
            r'קוד קבוצה:\s*([^\n]+)',
        ],
        'student1': [
            r'2\.\s*(?:Student one|Member A|שם חבר צוות ראשון):\s*([^\n]+)',
            r'Student one:\s*([^\n]+)',
            r'Member A:\s*([^\n]+)',
            r'Student 1:\s*([^\n]+)',
        ],
        'student2': [
            r'3\.\s*(?:Student two|Member B|שם חבר צוות שני):\s*([^\n]+)',
            r'Student two:\s*([^\n]+)',
            r'Member B:\s*([^\n]+)',
            r'Student 2:\s*([^\n]+)',
        ],
        'github': [
            r'4\.\s*(?:Repo link|GitHub Repository|קישור לריפו):\s*([^\n]+)',
            r'GitHub Repository:\s*([^\n]+)',
            r'Repo link:\s*([^\n]+)',
            r'(https://github\.com/[^\s\)]+)',
        ],
        'grade': [
            r'5\.\s*(?:Grade suggestion|הציון העצמי שלי):\s*(\d+)',
            r'Grade suggestion:\s*(\d+)',
            r'הציון העצמי שלי\s*[:]?\s*(\d+)',
            r'(?:Suggested [Gg]rade|Self [Gg]rade):\s*(\d+)',
        ]
    }

    # Try each pattern for each field
    for field, pattern_list in patterns.items():
        for pattern in pattern_list:
            match = re.search(pattern, pdf_text, re.IGNORECASE | re.MULTILINE)
            if match:
                info[field] = match.group(1).strip()
                break

    # Clean up GitHub URL - extract just the URL if there's extra text
    if info['github']:
        github_match = re.search(r'https://github\.com/[^\s\)]+', info['github'])
        if github_match:
            info['github'] = github_match.group(0)

    # Extract just the number from suggested grade
    if info['grade']:
        grade_match = re.search(r'(\d+)', info['grade'])
        if grade_match:
            info['grade'] = grade_match.group(1)

    return info


def create_student_excel(folder_path, participant_id, group_code, student1, student2, github, grade, pdf_filename):
    """Create an individual Excel file for a student submission."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission Info"

    # Headers
    headers = ["Field", "Value"]

    # Write headers with formatting
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    data = [
        ("Participant ID", participant_id),
        ("Group Code", group_code),
        ("Student 1", student1),
        ("Student 2", student2),
        ("GitHub Repository", github),
        ("Suggested Grade", grade),
        ("PDF Filename", pdf_filename)
    ]

    for row_idx, (field, value) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        cell = ws.cell(row=row_idx, column=2, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Make GitHub link clickable
        if field == "GitHub Repository" and value and value.startswith('http'):
            cell.hyperlink = value
            cell.font = Font(color="0563C1", underline="single")

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    # Save the Excel file in the student's folder
    output_path = folder_path / "submission_info.xlsx"
    wb.save(output_path)
    return output_path


def find_main_pdf(folder_path):
    """Find the main assignment PDF (not submission_keys.pdf)."""
    pdf_files = list(folder_path.glob('*.pdf'))
    # Filter out submission_keys.pdf
    main_pdfs = [p for p in pdf_files if 'submission_keys' not in p.name.lower()]
    return main_pdfs[0] if main_pdfs else (pdf_files[0] if pdf_files else None)


def main():
    """Process all participant folders in WorkSubmissions05."""
    base_path = Path('.')

    # Find all participant folders
    participant_folders = sorted([
        d for d in base_path.iterdir()
        if d.is_dir() and d.name.startswith('Participant_')
    ])

    print(f"Found {len(participant_folders)} participant folders in WorkSubmissions05\n")
    print("=" * 80)

    created_count = 0
    failed_count = 0
    missing_pdf_count = 0

    # Track which submissions need manual review
    needs_review = []

    for folder in participant_folders:
        participant_id = extract_participant_id(folder.name)
        pdf_file = find_main_pdf(folder)

        if not pdf_file:
            print(f"[X] {participant_id}: NO PDF FOUND")
            missing_pdf_count += 1
            continue

        print(f"\n[*] Processing {participant_id}: {pdf_file.name}")

        # This script outputs the info it found, and you (Claude) will use the Read tool
        # to read the PDF and then call create_student_excel with the extracted data
        print(f"    PDF Path: {pdf_file}")
        print(f"    Folder: {folder}")
        print(f"    Participant ID: {participant_id}")
        print(f"    Ready for Claude to read PDF and extract data...")

    print("\n" + "=" * 80)
    print(f"\nTotal participant folders: {len(participant_folders)}")
    print(f"PDFs ready to process: {len(participant_folders) - missing_pdf_count}")
    print(f"Missing PDFs: {missing_pdf_count}")
    print(f"\nNext steps:")
    print("1. Claude will read each PDF using the Read tool (first ~200 lines)")
    print("2. Extract the 7 required fields")
    print("3. Call create_student_excel() for each submission")
    print("4. Report completion status")


if __name__ == '__main__':
    # Check if openpyxl is installed
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is not installed")
        print("Please run: pip install openpyxl")
        sys.exit(1)

    main()
