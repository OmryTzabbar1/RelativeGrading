#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Process WorkSubmissions05 PDFs and create individual submission_info.xlsx files
Extracts information from first 2 pages and populates Excel in each folder
"""

import sys
import codecs
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Set UTF-8 encoding for stdout on Windows
if sys.platform == 'win32':
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')


def extract_participant_id(folder_name):
    """Extract participant ID from folder name."""
    match = re.search(r'Participant_(\d+)', folder_name)
    return match.group(1) if match else folder_name


def create_student_excel(folder_path, participant_id, group_code, student1, student2, github, grade, pdf_filename):
    """Create an individual Excel file for a student submission."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission Info"

    # Headers
    headers = ["Field", "Value"]

    # Write headers with formatting
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    data = [
        ("Participant ID", participant_id),
        ("Group Code", group_code),
        ("Student 1", student1),
        ("Student 2", student2),
        ("GitHub Repository", github),
        ("Suggested Grade", grade),
        ("PDF Filename", pdf_filename)
    ]

    for row_idx, (field, value) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        cell = ws.cell(row=row_idx, column=2, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Make GitHub link clickable
        if field == "GitHub Repository" and value and value.startswith('http'):
            cell.hyperlink = value
            cell.font = Font(color="0563C1", underline="single")

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    # Save the Excel file in the student's folder
    output_path = folder_path / "submission_info.xlsx"
    wb.save(output_path)
    return output_path


def find_main_pdf(folder_path):
    """Find the main assignment PDF (not submission_keys.pdf)."""
    pdf_files = list(folder_path.glob('*.pdf'))
    # Filter out submission_keys.pdf
    main_pdfs = [p for p in pdf_files if 'submission_keys' not in p.name.lower()]
    return main_pdfs[0] if main_pdfs else (pdf_files[0] if pdf_files else None)


def main():
    """List all participant folders and their PDFs for processing."""
    base_path = Path('.')

    # Find all participant folders
    participant_folders = sorted([
        d for d in base_path.iterdir()
        if d.is_dir() and d.name.startswith('Participant_')
    ])

    print(f"Found {len(participant_folders)} participant folders in WorkSubmissions05\n")
    print("=" * 80)

    submissions = []
    for folder in participant_folders:
        participant_id = extract_participant_id(folder.name)
        pdf_file = find_main_pdf(folder)

        if not pdf_file:
            print(f"[X] {participant_id}: NO PDF FOUND")
            continue

        submission = {
            'participant_id': participant_id,
            'folder': folder.name,
            'folder_path': folder,
            'pdf_path': str(pdf_file),
            'pdf_filename': pdf_file.name
        }
        submissions.append(submission)

        print(f"\n[*] Participant {participant_id}")
        print(f"    Folder: {folder.name}")
        print(f"    PDF: {pdf_file.name}")
        print(f"    Full path: {pdf_file}")

    print("\n" + "=" * 80)
    print(f"\nTotal submissions with PDFs: {len(submissions)}")
    print("\nNext: Claude will read each PDF and extract the required information")

    return submissions


if __name__ == '__main__':
    # Check if openpyxl is installed
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is not installed")
        print("Please run: pip install openpyxl")
        sys.exit(1)

    submissions = main()

    print("\n" + "=" * 80)
    print("Ready to process submissions!")
    print("Claude will now read each PDF and create Excel files...")
