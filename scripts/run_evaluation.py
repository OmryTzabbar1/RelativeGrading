#!/usr/bin/env python3
"""
Integrated Student Project Evaluator
Includes code verification and assignment profiles
"""

import sys
import os
import json
import re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Add skills directory to path
skills_dir = Path(__file__).parent / ".claude" / "skills" / "evaluating-student-projects"
sys.path.insert(0, str(skills_dir))

# Import the new modules
from code_analysis import run_full_code_analysis, format_criteria_summary
from assignment_profiles import detect_assignment_type, apply_assignment_profile

def find_markdown_files(student_folder):
    """Recursively find all .md files in student folder"""
    md_files = []
    for root, dirs, files in os.walk(student_folder):
        # Skip hidden directories
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        for file in files:
            if file.endswith('.md'):
                md_files.append(Path(root) / file)
    return md_files

def extract_criteria_from_markdown(md_content, filename):
    """
    Extract criteria from markdown content
    Returns list of criteria found
    """
    criteria = []

    # Filename-based detection
    filename_lower = filename.lower()
    filename_criteria = {
        'prd.md': 'PRD Document',
        'testing.md': 'Testing Documentation',
        'contributing.md': 'Contributing Guide',
        'architecture.md': 'Architecture Documentation',
        'changelog.md': 'Changelog',
        'api.md': 'API Documentation',
        'roadmap.md': 'Roadmap',
    }

    for pattern, criterion in filename_criteria.items():
        if pattern in filename_lower:
            criteria.append(criterion)

    # Content-based extraction patterns
    positive_patterns = [
        (r'we (built|created|developed|implemented)', 'implementation'),
        (r'the project includes?', 'feature'),
        (r'(unit|integration|e2e) tests?', 'testing'),
        (r'test coverage', 'testing'),
        (r'ci/cd', 'devops'),
        (r'docker', 'devops'),
        (r'eslint|pylint|ruff|black|prettier', 'code quality'),
        (r'pre-commit', 'code quality'),
        (r'typescript|mypy', 'type checking'),
        (r'screenshots?', 'visuals'),
        (r'problem statement', 'planning'),
        (r'cost analysis', 'research'),
        (r'requirements?', 'planning'),
        (r'success metrics', 'planning'),
        (r'architecture', 'documentation'),
    ]

    # Negative patterns (things to skip)
    negative_patterns = [
        r'todo:',
        r'future work',
        r'planned',
        r'not yet implemented',
        r'coming soon',
        r'stretch goal',
    ]

    lines = md_content.lower().split('\n')

    for i, line in enumerate(lines):
        # Skip negative contexts
        if any(re.search(neg, line, re.IGNORECASE) for neg in negative_patterns):
            continue

        # Extract from positive patterns
        if 'unit test' in line and 'test' not in [c.lower() for c in criteria]:
            criteria.append('Unit Tests')
        if 'integration test' in line:
            criteria.append('Integration Tests')
        if 'e2e test' in line or 'end-to-end test' in line:
            criteria.append('E2E Tests')
        if 'test coverage' in line:
            criteria.append('Test Coverage Metrics')
        if 'ci/cd' in line or 'github actions' in line:
            criteria.append('CI/CD Pipeline')
        if 'docker' in line:
            criteria.append('Docker Containerization')
        if 'eslint' in line:
            criteria.append('ESLint Configuration')
        if 'pylint' in line:
            criteria.append('Pylint Configuration')
        if 'ruff' in line:
            criteria.append('Ruff Linting')
        if 'black' in line and 'formatting' in line:
            criteria.append('Black Code Formatting')
        if 'prettier' in line:
            criteria.append('Prettier Configuration')
        if 'pre-commit' in line:
            criteria.append('Pre-commit Hooks')
        if 'typescript' in line:
            criteria.append('TypeScript Type Checking')
        if 'mypy' in line:
            criteria.append('MyPy Type Checking')
        if 'screenshot' in line:
            criteria.append('Screenshots')
        if 'problem statement' in line:
            criteria.append('Problem Statement')
        if 'cost analysis' in line or 'cost breakdown' in line:
            criteria.append('Cost Analysis')
        if 'requirements' in line and i > 0:
            criteria.append('Functional Requirements')
        if 'success metrics' in line or 'kpis' in line:
            criteria.append('Success Metrics')
        if 'architecture' in line and 'document' in line:
            criteria.append('Architecture Documentation')
        if 'readme' in line or filename_lower == 'readme.md':
            if 'README' not in criteria:
                criteria.append('README')

    return list(set(criteria))  # Remove duplicates

def categorize_criterion(criterion_name):
    """Categorize a criterion into broad topics"""
    criterion_lower = criterion_name.lower()

    # CodeQuality
    if any(kw in criterion_lower for kw in ['eslint', 'pylint', 'ruff', 'flake8', 'black', 'prettier',
                                              'pre-commit', 'typescript', 'mypy', 'type checking',
                                              'linting', 'formatting', 'code quality', 'verified']):
        return 'CodeQuality'

    # Testing
    if any(kw in criterion_lower for kw in ['test', 'coverage', 'pytest', 'jest']):
        return 'Testing'

    # DevOps
    if any(kw in criterion_lower for kw in ['ci/cd', 'docker', 'deployment', 'pipeline', '.gitignore']):
        return 'DevOps'

    # Documentation
    if any(kw in criterion_lower for kw in ['readme', 'documentation', 'api doc', 'changelog',
                                              'contributing', 'prompt_book']):
        return 'Documentation'

    # Planning
    if any(kw in criterion_lower for kw in ['prd', 'problem statement', 'requirements', 'architecture',
                                              'success metrics', 'assumptions']):
        return 'Planning'

    # Research
    if any(kw in criterion_lower for kw in ['cost analysis', 'risk analysis', 'roi', 'market research']):
        return 'Research'

    # Visuals
    if any(kw in criterion_lower for kw in ['screenshot', 'diagram', 'visualization']):
        return 'Visuals'

    # Security (new category for verified criteria)
    if any(kw in criterion_lower for kw in ['secret', 'security', '.env']):
        return 'Security'

    return 'Uncategorized'

def main():
    worksubmissions_folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("tests/WorkSubmissions04")

    print("\n" + "="*80)
    print("INTEGRATED STUDENT PROJECT EVALUATOR")
    print("With Code Verification & Assignment Profiles")
    print("="*80)

    # Step 2: Discover students
    student_folders = [f for f in worksubmissions_folder.iterdir()
                      if f.is_dir() and f.name.startswith('Participant_')]

    print(f"\n[Step 2/12] Found {len(student_folders)} student folders")

    # Data structures
    student_criteria = {}  # student_name -> [criteria]
    criteria_graph = {
        "metadata": {
            "total_students": len(student_folders),
            "total_criteria": 0,
            "assignment_folder": str(worksubmissions_folder)
        },
        "criteria": {}
    }

    # Step 3-4c: Read markdown, extract criteria, verify with code
    print(f"\n[Step 3-4c/12] Reading markdown files and verifying with code analysis...")

    for i, student_folder in enumerate(student_folders, 1):
        student_name = student_folder.name
        print(f"  [{i}/{len(student_folders)}] {student_name}...", end='', flush=True)

        # Step 3-4: Extract from markdown
        md_files = find_markdown_files(student_folder)
        student_crits = []

        for md_file in md_files:
            try:
                content = md_file.read_text(encoding='utf-8', errors='ignore')
                file_criteria = extract_criteria_from_markdown(content, md_file.name)
                student_crits.extend(file_criteria)
            except Exception as e:
                print(f"\n    Warning: Could not read {md_file}: {e}")

        # Step 4c: CODE VERIFICATION (NEW!)
        try:
            code_results = run_full_code_analysis(str(student_folder))
            verified_criteria = format_criteria_summary(code_results)
            student_crits.extend(verified_criteria)
            print(f" {len(verified_criteria)} verified", end='')
        except Exception as e:
            print(f" [code analysis failed: {e}]", end='')

        # Remove duplicates
        student_crits = list(set(student_crits))
        student_criteria[student_name] = student_crits

        print(f" -> {len(student_crits)} total criteria")

    # Step 5: Build criteria graph
    print(f"\n[Step 5/12] Building criteria graph...")
    all_criteria = set()
    for crits in student_criteria.values():
        all_criteria.update(crits)

    for criterion in all_criteria:
        students_with_criterion = [s for s, crits in student_criteria.items() if criterion in crits]
        category = categorize_criterion(criterion)

        criteria_graph["criteria"][criterion] = {
            "students": students_with_criterion,
            "count": len(students_with_criterion),
            "category": category,
            "weight": len(students_with_criterion) / len(student_folders)
        }

    criteria_graph["metadata"]["total_criteria"] = len(all_criteria)

    print(f"  Discovered {len(all_criteria)} unique criteria")

    # Step 7: Categorize summary
    print(f"\n[Step 7/12] Categorizing criteria...")
    category_counts = defaultdict(int)
    for criterion_data in criteria_graph["criteria"].values():
        category_counts[criterion_data["category"]] += 1

    for category, count in sorted(category_counts.items()):
        print(f"  {category}: {count} criteria")

    # Step 8b: APPLY ASSIGNMENT PROFILE (NEW!)
    print(f"\n[Step 8b/12] Applying assignment-specific calibration...")
    assignment_type = detect_assignment_type(worksubmissions_folder)

    if assignment_type:
        criteria_graph = apply_assignment_profile(criteria_graph, assignment_type)
    else:
        print("  No assignment profile detected, using default weights")

    # Step 9: Score and grade with rarity bonuses (NO CURVE)
    print(f"\n[Step 9/12] Calculating grades with rarity bonuses...")

    grades = []
    max_possible = sum(c["weight"] for c in criteria_graph["criteria"].values())
    total_students = len(student_criteria)

    # Identify rare criteria (≤15% prevalence)
    rare_criteria = set()
    for criterion, crit_data in criteria_graph["criteria"].items():
        prevalence = crit_data["count"] / total_students
        if prevalence <= 0.15:
            rare_criteria.add(criterion)

    print(f"  Rare criteria (<=15%%): {len(rare_criteria)}")

    for student_name, crits in student_criteria.items():
        # Base score
        score = sum(
            criteria_graph["criteria"][c]["weight"]
            for c in crits
            if c in criteria_graph["criteria"]
        )
        percentage = (score / max_possible * 100) if max_possible > 0 else 0

        # Calculate rarity bonus
        rare_count = sum(1 for c in crits if c in rare_criteria)
        rarity_bonus = rare_count * 1.0  # +1 point per rare criterion

        # Final grade = percentage + rarity bonus (capped at 100)
        final_grade = min(100.0, percentage + rarity_bonus)

        grades.append({
            "student": student_name,
            "score": score,
            "max_possible": max_possible,
            "percentage": percentage,
            "rarity_bonus": rarity_bonus,
            "grade": final_grade,
            "criteria_count": len(crits),
            "rare_criteria_count": rare_count
        })

    # Rank by final grade
    grades.sort(key=lambda x: x["grade"], reverse=True)
    for rank, grade_data in enumerate(grades, 1):
        grade_data["rank"] = rank

    print(f"  Top student: {grades[0]['student']} ({grades[0]['grade']:.1f})")
    print(f"  Average grade: {sum(g['grade'] for g in grades) / len(grades):.1f}")
    print(f"  Average rarity bonus: +{sum(g['rarity_bonus'] for g in grades) / len(grades):.2f}")

    # Step 10: Generate outputs
    print(f"\n[Step 10/12] Generating output files...")

    output_dir = Path("outputs")
    output_dir.mkdir(exist_ok=True)

    # Save criteria graph
    with open(output_dir / "criteria_graph_final.json", 'w') as f:
        json.dump(criteria_graph, f, indent=2)
    print("  [OK] Saved criteria_graph_final.json")

    # Save grades Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades"

    # Headers - Include rarity bonus
    headers = ["Student", "Raw Score", "Max Possible", "Percentage", "Rarity Bonus", "Grade", "Rank", "Criteria Count"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Data - Include rarity bonus column
    for row, grade_data in enumerate(grades, 2):
        ws.cell(row, 1, grade_data["student"])           # Student ID
        ws.cell(row, 2, grade_data['score'])             # Raw Score (number)
        ws.cell(row, 3, grade_data['max_possible'])      # Max Possible (number)
        ws.cell(row, 4, grade_data['percentage'])        # Percentage (number)
        ws.cell(row, 5, grade_data['rarity_bonus'])      # Rarity Bonus (NEW!)
        ws.cell(row, 6, grade_data['grade'])             # Grade (percentage + bonus, capped at 100)
        ws.cell(row, 7, grade_data["rank"])              # Rank
        ws.cell(row, 8, grade_data["criteria_count"])    # Criteria Count

    wb.save(output_dir / "grades.xlsx")
    print("  [OK] Saved grades.xlsx")

    # Summary report
    with open(output_dir / "EVALUATION_SUMMARY.md", 'w') as f:
        f.write("# Evaluation Summary\n\n")
        f.write(f"**Students evaluated:** {len(student_folders)}\n")
        f.write(f"**Criteria discovered:** {len(all_criteria)}\n")
        f.write(f"**Verified via code analysis:** {sum(1 for c in all_criteria if '(verified)' in c)}\n\n")

        if assignment_type:
            profile_meta = criteria_graph["metadata"].get("assignment_profile", {})
            f.write(f"**Assignment Profile:** {profile_meta.get('name', assignment_type)}\n")
            f.write(f"**Focus areas:** {', '.join(profile_meta.get('focus_areas', []))}\n\n")

        f.write("## Top 5 Students\n\n")
        for i, grade_data in enumerate(grades[:5], 1):
            f.write(f"{i}. {grade_data['student']} ({grade_data['grade']:.1f}) - {grade_data['criteria_count']} criteria\n")

        f.write("\n## Criteria by Category\n\n")
        for category, count in sorted(category_counts.items()):
            f.write(f"- **{category}**: {count} criteria\n")

    print("  [OK] Saved EVALUATION_SUMMARY.md")

    print("\n" + "="*80)
    print("EVALUATION COMPLETE!")
    print("="*80)
    print(f"\nOutput files saved to: {output_dir.absolute()}")
    print("\nNext steps:")
    print(f"  1. python organize_outputs.py {worksubmissions_folder.name}")
    print(f"  2. python compare_grades.py {worksubmissions_folder.name}")

if __name__ == "__main__":
    main()
