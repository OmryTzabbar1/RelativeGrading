#!/usr/bin/env python3
"""
Evaluate student projects in WorkSubmissions04
Following the complete workflow from SKILL.md
"""

import os
import re
import json
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Base paths
BASE_DIR = Path(r"E:\Projects\student-project-evaluator\tests\WorkSubmissions04")
OUTPUT_DIR = Path(r"E:\Projects\student-project-evaluator\outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

# Category keywords from CATEGORIES.md
CATEGORIES = {
    "Documentation": ["readme", "api docs", "api documentation", "user guide", "usage guide",
                     "changelog", "change log", "contributing", "contribution guide", "license",
                     "installation guide", "setup guide", "getting started", "faq", "wiki", "manual"],

    "Planning": ["prd", "product requirements", "architecture", "design document", "design doc",
                "technical spec", "roadmap", "requirements", "functional requirements", "non-functional requirements",
                "use cases", "use case", "problem statement", "problem", "solution overview", "proposed solution",
                "project goals", "objectives", "success metrics", "kpis", "assumptions", "constraints",
                "limitations", "scope", "design decisions", "trade-offs", "user personas", "user flow",
                "timeline", "schedule", "project plan"],

    "Testing": ["unit test", "integration test", "e2e test", "end-to-end", "test coverage",
               "code coverage", "pytest", "jest", "mocha", "junit", "testing framework",
               "test suite", "automated tests", "test cases"],

    "DevOps": ["ci/cd", "ci cd", "continuous integration", "github actions", "docker",
              "dockerfile", "containerization", "kubernetes", "deployment", "aws", "azure",
              "gcp", "cloud", "monitoring", "logging"],

    "Research": ["research", "analysis", "data analysis", "jupyter notebook", "notebook",
                "experiment", "findings", "benchmarking", "survey", "user research"],

    "Visuals": ["screenshot", "diagram", "flowchart", "chart", "demo video", "gif",
               "mockup", "wireframe", "architecture diagram", "sequence diagram"],

    "CodeQuality": ["linting", "linter", "eslint", "pylint", "ruff", "flake8", "prettier",
                   "black", "autopep8", "type checking", "typescript", "mypy", "pre-commit",
                   "git hooks", "code review", "code style", "style guide", "pep 8", "pep8",
                   "static analysis", "sonarqube", "quality gates", "iso 25010", "iso25010"],

    "Business": ["cost analysis", "budget", "roi", "return on investment", "market research",
                "business case", "competitive analysis", "risk analysis", "risk assessment"]
}

# Filename-based criteria detection
FILENAME_CRITERIA = {
    r"PRD\.md|prd\.md|ProductRequirements\.md": "PRD Document",
    r"TESTING\.md|testing\.md|TEST\.md": "Testing Documentation",
    r"CONTRIBUTING\.md|contributing\.md": "Contributing Guide",
    r"QUICKSTART\.md|QuickStart\.md|quick-start\.md": "Quick Start Guide",
    r"ARCHITECTURE\.md|architecture\.md|DESIGN\.md": "Architecture Documentation",
    r"CHANGELOG\.md|changelog\.md|CHANGES\.md": "Changelog",
    r"API\.md|api\.md|API_DOCS\.md": "API Documentation",
    r"ROADMAP\.md|roadmap\.md": "Roadmap",
    r"DEPLOYMENT\.md|deployment\.md|DEPLOY\.md": "Deployment Guide",
    r"TROUBLESHOOTING\.md|troubleshooting\.md|FAQ\.md": "Troubleshooting Guide",
}

# Positive implementation indicators (case-insensitive)
IMPLEMENTATION_INDICATORS = [
    r"\bwe built\b", r"\bwe created\b", r"\bwe developed\b", r"\bwe implemented\b",
    r"\bproject includes\b", r"\bthis project includes\b", r"\bincludes\b",
    r"\bfeatures?:\b", r"\bkey features\b", r"\bimplemented\b",
    r"\bcreated\b", r"\bdeveloped\b", r"\bbuilt\b", r"\badded\b",
    r"\bintegrated\b", r"\bsupports\b", r"\bprovides\b", r"\boffers\b",
    r"\benables\b", r"\bhas\b", r"\bcontains\b", r"\bcomes with\b",
    r"\bcomplete\b", r"\bcomprehensive\b", r"\bavailable\b",
    r"\btotal tests\b", r"\btest cases\b", r"\bcoverage\b",
]

# Negative indicators (skip these)
NEGATIVE_INDICATORS = [
    r"\bTODO\b", r"\bFIXME\b", r"\bwill add\b", r"\bplanning to\b",
    r"\bplanned for\b", r"\bnext steps\b", r"\bnot yet implemented\b",
    r"\bnot complete\b", r"\bmissing\b", r"\bout of scope\b",
    r"\bnot working\b", r"\bskipped\b"
]

def normalize_criterion_name(name):
    """Normalize criterion names for consistency"""
    # Remove extra whitespace
    name = " ".join(name.split())

    # Common normalizations
    normalizations = {
        r"\bunit test(s|ing)?\b": "Unit Tests",
        r"\bintegration test(s|ing)?\b": "Integration Tests",
        r"\be2e test(s)?\b": "E2E Tests",
        r"\bend-to-end test(s)?\b": "E2E Tests",
        r"\btest coverage\b": "Test Coverage Metrics",
        r"\bcode coverage\b": "Test Coverage Metrics",
        r"\breadme(\.md)?\b": "README",
        r"\bci/cd\b": "CI/CD Pipeline",
        r"\bdocker(file)?\b": "Docker",
        r"\bapi doc(s|umentation)?\b": "API Documentation",
        r"\barchitecture doc(s|umentation)?\b": "Architecture Documentation",
        r"\bprd doc(ument)?\b": "PRD Document",
    }

    name_lower = name.lower()
    for pattern, replacement in normalizations.items():
        if re.search(pattern, name_lower, re.IGNORECASE):
            return replacement

    # Title case the result
    return " ".join(word.capitalize() for word in name.split())

def has_negative_context(text, match_pos):
    """Check if match is in a negative context"""
    # Get surrounding context (200 chars before and after)
    start = max(0, match_pos - 200)
    end = min(len(text), match_pos + 200)
    context = text[start:end].lower()

    for neg_pattern in NEGATIVE_INDICATORS:
        if re.search(neg_pattern, context, re.IGNORECASE):
            return True
    return False

def extract_criteria_from_file(filepath):
    """Extract criteria from a single markdown file"""
    criteria = set()

    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

        filename = os.path.basename(filepath)

        # 1. Check filename-based criteria
        for pattern, criterion in FILENAME_CRITERIA.items():
            if re.search(pattern, filename, re.IGNORECASE):
                criteria.add(criterion)

        # 2. Section header detection
        section_patterns = {
            r"^#{1,3}\s*(Testing|Tests)\b": "Testing Documentation",
            r"^#{1,3}\s*Unit Tests?\b": "Unit Tests",
            r"^#{1,3}\s*(Installation|Setup|Getting Started)\b": "Installation Instructions",
            r"^#{1,3}\s*(Usage|How to Use)\b": "Usage Guide",
            r"^#{1,3}\s*Features?\b": "Features Documentation",
            r"^#{1,3}\s*(Screenshots?|Demo|Examples?)\b": "Screenshots/Visuals",
            r"^#{1,3}\s*(Architecture|Design)\b": "Architecture Documentation",
            r"^#{1,3}\s*(CI/CD|Deployment|DevOps)\b": "DevOps Documentation",
            r"^#{1,3}\s*(API|Endpoints|API Reference)\b": "API Documentation",
            r"^#{1,3}\s*Contributing\b": "Contributing Guide",
            r"^#{1,3}\s*Problem Statement\b": "Problem Statement",
            r"^#{1,3}\s*(Solution|Proposed Solution)\b": "Solution Overview",
            r"^#{1,3}\s*Requirements?\b": "Requirements Documentation",
            r"^#{1,3}\s*Use Cases?\b": "Use Case Documentation",
            r"^#{1,3}\s*(Cost Analysis|Budget|Costs)\b": "Cost Analysis",
            r"^#{1,3}\s*(Assumptions|Constraints)\b": "Assumptions Documentation",
            r"^#{1,3}\s*(Success Metrics|KPIs|Metrics)\b": "Success Metrics",
            r"^#{1,3}\s*(Risk Analysis|Risks)\b": "Risk Analysis",
            r"^#{1,3}\s*(Roadmap|Future Work|Next Steps)\b": "Roadmap",
        }

        for line in content.split('\n'):
            for pattern, criterion in section_patterns.items():
                if re.search(pattern, line, re.IGNORECASE | re.MULTILINE):
                    # Check it's not in a TODO context
                    if not re.search(r'\bTODO\b', line, re.IGNORECASE):
                        criteria.add(criterion)

        # 3. Look for implementation indicators with specific features
        content_lower = content.lower()

        # Testing indicators
        if re.search(r'\b(unit test|test suite|pytest|jest)\b', content_lower):
            if not has_negative_context(content, content_lower.find('test')):
                criteria.add("Unit Tests")

        if re.search(r'\bintegration test', content_lower):
            criteria.add("Integration Tests")

        if re.search(r'\b(e2e|end-to-end) test', content_lower):
            criteria.add("E2E Tests")

        if re.search(r'\b(\d+%|coverage)', content_lower):
            if 'test' in content_lower or 'coverage' in content_lower:
                criteria.add("Test Coverage Metrics")

        # Quality Standards
        if re.search(r'\beslint\b', content_lower):
            criteria.add("ESLint Configuration")

        if re.search(r'\bpylint\b', content_lower):
            criteria.add("Pylint Configuration")

        if re.search(r'\bprettier\b', content_lower):
            criteria.add("Prettier Formatting")

        if re.search(r'\bblack\b', content_lower) and 'formatter' in content_lower:
            criteria.add("Black Formatting")

        if re.search(r'\bpre-commit\b', content_lower):
            criteria.add("Pre-commit Hooks")

        if re.search(r'\bmypy\b', content_lower):
            criteria.add("Mypy Type Checking")

        if re.search(r'\btypescript\b', content_lower) and 'strict' in content_lower:
            criteria.add("TypeScript Type Checking")

        if re.search(r'\bcode review\b', content_lower):
            criteria.add("Code Review Process")

        if re.search(r'\b(pep ?8|pep8)\b', content_lower):
            criteria.add("PEP8 Compliance")

        if re.search(r'\bruff\b', content_lower):
            criteria.add("Ruff Linting")

        if re.search(r'\b(sonarqube|sonarcloud)\b', content_lower):
            criteria.add("Static Analysis Tools")

        if re.search(r'\bquality gate', content_lower):
            criteria.add("Quality Gates")

        if re.search(r'\b(code style|coding standard)', content_lower):
            criteria.add("Code Style Guide")

        if re.search(r'\biso.?25010\b', content_lower):
            criteria.add("ISO 25010 Compliance")

        # DevOps
        if re.search(r'\b(ci/cd|github actions|gitlab ci)\b', content_lower):
            criteria.add("CI/CD Pipeline")

        if re.search(r'\bdocker', content_lower):
            criteria.add("Docker")

        # Documentation
        if 'readme' in filename.lower():
            criteria.add("README")

        # Additional specific documentation
        if re.search(r'\bcoverage report\b', content_lower):
            criteria.add("Coverage Report")

        if re.search(r'\bexecution log\b', content_lower):
            criteria.add("Execution Log")

        if re.search(r'\bextensibility\b', content_lower) and '##' in content:
            criteria.add("Extensibility Documentation")

        if re.search(r'\bparameter analysis\b', content_lower):
            criteria.add("Parameter Analysis")

        if re.search(r'\bai model selection\b', content_lower):
            criteria.add("AI Model Selection Documentation")

        if re.search(r'\bprompt engineering\b', content_lower) and 'log' in content_lower:
            criteria.add("Prompt Engineering Log")

        if re.search(r'\bsecurity\b', content_lower) and ('policy' in content_lower or 'guideline' in content_lower):
            criteria.add("Security Documentation")

    except Exception as e:
        print(f"Error processing {filepath}: {e}")

    return criteria

def extract_student_id(folder_name):
    """Extract student ID from folder name like Participant_87681_assignsubmission_file"""
    match = re.search(r'Participant_(\d+)_', folder_name)
    if match:
        return match.group(1)
    return folder_name

def categorize_criterion(criterion_name):
    """Categorize a criterion based on keywords"""
    criterion_lower = criterion_name.lower()

    for category, keywords in CATEGORIES.items():
        for keyword in keywords:
            if keyword.lower() in criterion_lower:
                return category

    return "Uncategorized"

def main():
    print("=" * 80)
    print("STUDENT PROJECT EVALUATION - WorkSubmissions04")
    print("=" * 80)
    print()

    # STEP 1-2: Discover students
    print("STEP 1-2: Discovering students...")
    student_folders = sorted([d for d in os.listdir(BASE_DIR)
                            if d.startswith("Participant_") and
                            os.path.isdir(BASE_DIR / d)])

    total_students = len(student_folders)
    print(f"Found {total_students} student folders")
    print()

    # STEP 3-4: Read and Extract
    print("STEP 3-4: Reading markdown files and extracting criteria...")
    student_criteria = {}  # {student_id: set of criteria}
    criteria_sources = defaultdict(list)  # {criterion: [student_ids]}

    for folder in student_folders:
        student_id = extract_student_id(folder)
        student_path = BASE_DIR / folder

        print(f"Processing {student_id}...", end=" ")

        # Find all markdown files
        md_files = list(student_path.rglob("*.md"))
        print(f"({len(md_files)} markdown files)")

        # Extract criteria from all markdown files
        all_criteria = set()
        for md_file in md_files:
            criteria = extract_criteria_from_file(md_file)
            all_criteria.update(criteria)

        student_criteria[student_id] = all_criteria

        # Track which students have which criteria
        for criterion in all_criteria:
            criteria_sources[criterion].append(student_id)

    print()
    print(f"Total unique criteria discovered: {len(criteria_sources)}")
    print()

    # STEP 5: Build Criteria Graph
    print("STEP 5: Building criteria graph...")
    criteria_graph = {
        "metadata": {
            "total_students": total_students,
            "evaluation_date": datetime.now().strftime("%Y-%m-%d"),
            "total_criteria": len(criteria_sources)
        },
        "criteria": {}
    }

    for criterion, students in criteria_sources.items():
        count = len(students)
        weight = count / total_students

        criteria_graph["criteria"][criterion] = {
            "display_name": criterion,
            "students": sorted(students),
            "count": count,
            "weight": weight,
            "category": "Uncategorized"  # Will be set in next step
        }

    # STEP 6: Categorize
    print("STEP 6: Categorizing criteria...")
    for criterion in criteria_graph["criteria"]:
        category = categorize_criterion(criterion)
        criteria_graph["criteria"][criterion]["category"] = category

    # Count by category
    category_counts = defaultdict(int)
    for criterion_data in criteria_graph["criteria"].values():
        category_counts[criterion_data["category"]] += 1

    print("Criteria by category:")
    for category, count in sorted(category_counts.items()):
        print(f"  {category}: {count}")
    print()

    # STEP 7: Calculate Weights with Rarity Bonus
    print("STEP 7: Calculating weights with rarity bonus...")
    for criterion, data in criteria_graph["criteria"].items():
        base_weight = data["count"] / total_students

        # Rarity bonus: ≤15% prevalence (≤3 students out of 20) gets +1 bonus
        rarity_threshold = total_students * 0.15
        if data["count"] <= rarity_threshold:
            bonus = 1.0
            data["weight"] = base_weight + bonus
            data["has_rarity_bonus"] = True
        else:
            data["weight"] = base_weight
            data["has_rarity_bonus"] = False

    # STEP 8: Score and Grade
    print("STEP 8: Calculating scores and grades...")

    # Calculate max possible score
    max_possible = sum(data["weight"] for data in criteria_graph["criteria"].values())

    student_scores = []
    for student_id, criteria_set in student_criteria.items():
        # Calculate raw score
        raw_score = sum(criteria_graph["criteria"][c]["weight"]
                       for c in criteria_set
                       if c in criteria_graph["criteria"])

        # Calculate percentage
        percentage = (raw_score / max_possible * 100) if max_possible > 0 else 0

        student_scores.append({
            "student_id": student_id,
            "raw_score": raw_score,
            "max_possible": max_possible,
            "percentage": percentage,
            "criteria_count": len(criteria_set),
            "criteria": sorted(list(criteria_set))
        })

    # Sort by percentage descending
    student_scores.sort(key=lambda x: x["percentage"], reverse=True)

    # Find best percentage for relative grading
    best_percentage = student_scores[0]["percentage"] if student_scores else 100

    # Assign ranks and relative grades
    for rank, student in enumerate(student_scores, 1):
        student["rank"] = rank
        student["relative_grade"] = (student["percentage"] / best_percentage * 100) if best_percentage > 0 else 0

    # STEP 9: Generate Outputs
    print("STEP 9: Generating output files...")

    # 1. Save criteria_graph_final.json
    graph_file = OUTPUT_DIR / "criteria_graph_final.json"
    with open(graph_file, 'w', encoding='utf-8') as f:
        json.dump(criteria_graph, f, indent=2, ensure_ascii=False)
    print(f"  [OK] Saved {graph_file}")

    # 2. Generate grades.xlsx
    grades_file = OUTPUT_DIR / "grades.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades"

    # Headers
    headers = ["Student ID", "Raw Score", "Max Possible", "Percentage", "Relative Grade", "Rank", "Criteria Count"]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for student in student_scores:
        ws.append([
            student["student_id"],
            round(student["raw_score"], 2),
            round(student["max_possible"], 2),
            round(student["percentage"], 2),
            round(student["relative_grade"], 2),
            student["rank"],
            student["criteria_count"]
        ])

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(grades_file)
    print(f"  [OK] Saved {grades_file}")

    # 3. Generate Student_Evaluation_Report.xlsx (multi-sheet)
    report_file = OUTPUT_DIR / "Student_Evaluation_Report.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["WorkSubmissions04 Evaluation Summary"])
    ws_summary.append([])
    ws_summary.append(["Total Students", total_students])
    ws_summary.append(["Total Criteria", len(criteria_sources)])
    ws_summary.append(["Evaluation Date", datetime.now().strftime("%Y-%m-%d")])
    ws_summary.append(["Max Possible Score", round(max_possible, 2)])
    ws_summary.append([])
    ws_summary.append(["Category", "Criteria Count"])
    for category, count in sorted(category_counts.items()):
        ws_summary.append([category, count])

    # Sheet 2: Student Grades (same as grades.xlsx)
    ws_grades = wb.create_sheet("Student Grades")
    ws_grades.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws_grades.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for student in student_scores:
        ws_grades.append([
            student["student_id"],
            round(student["raw_score"], 2),
            round(student["max_possible"], 2),
            round(student["percentage"], 2),
            round(student["relative_grade"], 2),
            student["rank"],
            student["criteria_count"]
        ])

    # Sheet 3: Criteria Details
    ws_criteria = wb.create_sheet("Criteria Details")
    criteria_headers = ["Criterion", "Category", "Student Count", "Weight", "Rarity Bonus"]
    ws_criteria.append(criteria_headers)

    for col in range(1, len(criteria_headers) + 1):
        cell = ws_criteria.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for criterion, data in sorted(criteria_graph["criteria"].items()):
        ws_criteria.append([
            criterion,
            data["category"],
            data["count"],
            round(data["weight"], 3),
            "Yes" if data.get("has_rarity_bonus", False) else "No"
        ])

    # Sheet 4: Student Details (each student's criteria)
    ws_details = wb.create_sheet("Student Details")
    ws_details.append(["Student ID", "Rank", "Criteria Count", "Criteria List"])

    for student in student_scores:
        ws_details.append([
            student["student_id"],
            student["rank"],
            student["criteria_count"],
            ", ".join(student["criteria"])
        ])

    wb.save(report_file)
    print(f"  [OK] Saved {report_file}")

    print()
    print("=" * 80)
    print("EVALUATION COMPLETE")
    print("=" * 80)
    print()
    print(f"Total Students: {total_students}")
    print(f"Total Criteria: {len(criteria_sources)}")
    print(f"CodeQuality Criteria: {category_counts['CodeQuality']}")
    print()
    print("Top 3 Students:")
    for i, student in enumerate(student_scores[:3], 1):
        print(f"  {i}. Student {student['student_id']}: "
              f"{student['percentage']:.1f}% "
              f"({student['criteria_count']} criteria)")
    print()

    # Grade distribution
    print("Grade Distribution:")
    ranges = [(90, 100, "A"), (80, 90, "B"), (70, 80, "C"), (60, 70, "D"), (0, 60, "F")]
    for low, high, grade in ranges:
        count = sum(1 for s in student_scores if low <= s["percentage"] < high)
        print(f"  {grade} ({low}-{high}%): {count} students")
    print()

    print("Output files:")
    print(f"  - {graph_file}")
    print(f"  - {grades_file}")
    print(f"  - {report_file}")
    print()

if __name__ == "__main__":
    main()
