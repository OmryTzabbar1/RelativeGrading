"""
Create comprehensive Excel report with student evaluation results
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from collections import Counter

# Load data
with open('outputs/criteria_graph_grouped.json', 'r') as f:
    grouped_data = json.load(f)

with open('outputs/criteria_graph_final.json', 'r') as f:
    original_data = json.load(f)

# Load grades
import openpyxl
grades_wb = openpyxl.load_workbook('outputs/grades.xlsx')
grades_ws = grades_wb['Grades']

# Create new workbook
wb = Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
subheader_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
top_fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
good_fill = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
poor_fill = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# ============================================================================
# SHEET 1: FINAL GRADES
# ============================================================================
ws1 = wb.active
ws1.title = "Final Grades"

headers = ["Rank", "Student ID", "Final Grade", "Base %", "Rarity Bonus",
           "Criteria Count", "Total Criteria", "Percentile"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Add data
for row in range(2, 38):
    rank = grades_ws.cell(row, 1).value
    student = grades_ws.cell(row, 2).value
    base = grades_ws.cell(row, 3).value
    bonus = grades_ws.cell(row, 4).value
    final = grades_ws.cell(row, 5).value
    criteria_str = grades_ws.cell(row, 6).value
    criteria_count = int(criteria_str.split('/')[0])
    total_criteria = int(criteria_str.split('/')[1])
    percentile = ((37 - rank) / 36) * 100

    values = [rank, student, final, base, bonus, criteria_count, total_criteria, percentile]

    for col, val in enumerate(values, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Number formatting
        if col in [3, 4, 8]:  # Final, Base, Percentile
            cell.number_format = '0.0'
        elif col == 5:  # Bonus
            cell.number_format = '+0;-0;0'

    # Color coding by grade
    if final >= 90:
        for c in range(1, 9):
            ws1.cell(row=row, column=c).fill = top_fill
    elif final >= 70:
        for c in range(1, 9):
            ws1.cell(row=row, column=c).fill = good_fill
    elif final < 50:
        for c in range(1, 9):
            ws1.cell(row=row, column=c).fill = poor_fill

# Set column widths
ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 14
ws1.column_dimensions['G'].width = 14
ws1.column_dimensions['H'].width = 12

ws1.freeze_panes = 'A2'

# ============================================================================
# SHEET 2: GRADE DISTRIBUTION
# ============================================================================
ws2 = wb.create_sheet("Grade Distribution")

# Headers
ws2['A1'] = "Grade Range"
ws2['B1'] = "Count"
ws2['C1'] = "Percentage"
ws2['D1'] = "Students"

for col in ['A', 'B', 'C', 'D']:
    ws2[f'{col}1'].font = header_font
    ws2[f'{col}1'].fill = header_fill
    ws2[f'{col}1'].alignment = Alignment(horizontal='center')
    ws2[f'{col}1'].border = thin_border

# Calculate distribution
grade_ranges = [
    ("90-100 (Excellent)", 90, 100),
    ("80-89 (Very Good)", 80, 89),
    ("70-79 (Good)", 70, 79),
    ("60-69 (Satisfactory)", 60, 69),
    ("50-59 (Pass)", 50, 59),
    ("40-49 (Below Average)", 40, 49),
    ("30-39 (Poor)", 30, 39),
    ("20-29 (Very Poor)", 20, 29),
    ("10-19 (Failing)", 10, 19)
]

distribution = []
for range_name, min_grade, max_grade in grade_ranges:
    students_in_range = []
    for row in range(2, 38):
        final = grades_ws.cell(row, 5).value
        if min_grade <= final <= max_grade:
            students_in_range.append(grades_ws.cell(row, 2).value)

    distribution.append((range_name, len(students_in_range), students_in_range))

row_num = 2
for range_name, count, students in distribution:
    ws2.cell(row=row_num, column=1, value=range_name).border = thin_border
    ws2.cell(row=row_num, column=2, value=count).border = thin_border
    ws2.cell(row=row_num, column=3, value=count/36*100).border = thin_border
    ws2.cell(row=row_num, column=3).number_format = '0.0"%"'
    ws2.cell(row=row_num, column=4, value=", ".join(students)).border = thin_border
    row_num += 1

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 80

# ============================================================================
# SHEET 3: CRITERIA BREAKDOWN (Top 10 Students)
# ============================================================================
ws3 = wb.create_sheet("Top Students Criteria")

ws3['A1'] = "Top 10 Students - Criteria Breakdown"
ws3['A1'].font = Font(bold=True, size=14)
ws3.merge_cells('A1:D1')

row_num = 3
for rank in range(1, 11):
    student = grades_ws.cell(rank + 1, 2).value
    final = grades_ws.cell(rank + 1, 5).value

    # Get criteria for this student
    student_criteria = []
    for crit_key, crit_data in grouped_data['criteria'].items():
        if student in crit_data['students']:
            student_criteria.append(crit_data['display_name'])

    # Header for this student
    ws3.cell(row=row_num, column=1, value=f"Rank {rank}: Student {student} (Grade: {final:.1f})")
    ws3.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    ws3.cell(row=row_num, column=1).fill = subheader_fill
    ws3.cell(row=row_num, column=1).font = Font(bold=True, color="FFFFFF")
    ws3.merge_cells(f'A{row_num}:D{row_num}')
    row_num += 1

    # Criteria (3 per row)
    for i in range(0, len(sorted(student_criteria)), 3):
        criteria_row = sorted(student_criteria)[i:i+3]
        for col, crit in enumerate(criteria_row, 1):
            ws3.cell(row=row_num, column=col, value=crit)
        row_num += 1

    row_num += 1  # Blank row

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 35

# ============================================================================
# SHEET 4: ALL CRITERIA LIST
# ============================================================================
ws4 = wb.create_sheet("All Criteria")

headers = ["#", "Criterion Name", "Students", "Prevalence %", "Category",
           "Specific Criteria (if grouped)"]
for col, header in enumerate(headers, 1):
    cell = ws4.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Sort criteria by prevalence
criteria_list = []
for crit_key, crit_data in grouped_data['criteria'].items():
    criteria_list.append((
        crit_data['display_name'],
        crit_data['count'],
        crit_data['count'] / 36 * 100,
        crit_data.get('category', 'Uncategorized'),
        crit_data.get('specific_criteria', [])
    ))

criteria_list.sort(key=lambda x: x[1], reverse=True)

for idx, (name, count, prevalence, category, specifics) in enumerate(criteria_list, 1):
    row = idx + 1

    specific_str = ", ".join(specifics) if len(specifics) > 1 else "-"

    values = [idx, name, count, prevalence, category, specific_str]
    for col, val in enumerate(values, 1):
        cell = ws4.cell(row=row, column=col, value=val)
        cell.border = thin_border
        if col == 4:
            cell.number_format = '0.0"%"'
        if col in [1, 3, 4]:
            cell.alignment = Alignment(horizontal='center')

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 35
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 18
ws4.column_dimensions['F'].width = 60

ws4.freeze_panes = 'A2'

# ============================================================================
# SHEET 5: COMPARISON (Before vs After Grouping)
# ============================================================================
ws5 = wb.create_sheet("Grouping Impact")

headers = ["Student ID", "Original Criteria", "Grouped Criteria", "Change", "% Change"]
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Calculate counts
def get_student_counts(criteria_dict):
    counts = {}
    for crit_data in criteria_dict.values():
        for student in crit_data['students']:
            counts[student] = counts.get(student, 0) + 1
    return counts

original_counts = get_student_counts(original_data['criteria'])
grouped_counts = get_student_counts(grouped_data['criteria'])

# Get all students sorted by original count
all_students = sorted(original_counts.keys(), key=lambda x: original_counts[x], reverse=True)

for idx, student in enumerate(all_students, 2):
    orig = original_counts[student]
    grp = grouped_counts.get(student, 0)
    change = grp - orig
    pct_change = (change / orig * 100) if orig > 0 else 0

    values = [student, orig, grp, change, pct_change]
    for col, val in enumerate(values, 1):
        cell = ws5.cell(row=idx, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

        if col == 4:  # Change
            cell.number_format = '+0;-0;0'
        elif col == 5:  # % Change
            cell.number_format = '+0.0%;-0.0%;0%'

        # Color coding
        if col >= 4 and change < 0:
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

ws5.column_dimensions['A'].width = 12
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 12
ws5.column_dimensions['E'].width = 12

# ============================================================================
# SHEET 6: SUMMARY STATISTICS
# ============================================================================
ws6 = wb.create_sheet("Summary")

ws6['A1'] = "EVALUATION SUMMARY STATISTICS"
ws6['A1'].font = Font(bold=True, size=16)
ws6.merge_cells('A1:B1')

stats = [
    ("", ""),
    ("Total Students Evaluated", 36),
    ("Total Criteria (Original)", 88),
    ("Total Criteria (Grouped)", 65),
    ("Criteria Reduction", "23 (26%)"),
    ("", ""),
    ("GRADES", ""),
    ("Highest Grade", f"{max(grades_ws.cell(row, 5).value for row in range(2, 38)):.1f}"),
    ("Lowest Grade", f"{min(grades_ws.cell(row, 5).value for row in range(2, 38)):.1f}"),
    ("Average Grade", f"{sum(grades_ws.cell(row, 5).value for row in range(2, 38))/36:.1f}"),
    ("Median Grade", "~60"),
    ("", ""),
    ("CRITERIA PER STUDENT", ""),
    ("Maximum", f"{max(grouped_counts.values())} criteria"),
    ("Minimum", f"{min(grouped_counts.values())} criteria"),
    ("Average", f"{sum(grouped_counts.values())/36:.1f} criteria"),
    ("", ""),
    ("TOP PERFORMERS", ""),
    ("Grade 90+", f"{sum(1 for row in range(2, 38) if grades_ws.cell(row, 5).value >= 90)} students"),
    ("Grade 80-89", f"{sum(1 for row in range(2, 38) if 80 <= grades_ws.cell(row, 5).value < 90)} students"),
    ("Grade 70-79", f"{sum(1 for row in range(2, 38) if 70 <= grades_ws.cell(row, 5).value < 80)} students"),
]

for idx, (label, value) in enumerate(stats, 2):
    if label == "":
        continue

    ws6.cell(row=idx, column=1, value=label)
    ws6.cell(row=idx, column=2, value=value)

    if label in ["GRADES", "CRITERIA PER STUDENT", "TOP PERFORMERS"]:
        ws6.cell(row=idx, column=1).font = Font(bold=True, size=12)
        ws6.cell(row=idx, column=1).fill = subheader_fill
        ws6.cell(row=idx, column=1).font = Font(bold=True, color="FFFFFF")
    else:
        ws6.cell(row=idx, column=1).font = Font(bold=True)

ws6.column_dimensions['A'].width = 35
ws6.column_dimensions['B'].width = 20

# Save
output_file = 'outputs/Student_Evaluation_Report.xlsx'
wb.save(output_file)
print(f"\nComprehensive Excel report saved: {output_file}")
print(f"\nSheets included:")
print("  1. Final Grades - Complete ranked list")
print("  2. Grade Distribution - Breakdown by grade ranges")
print("  3. Top Students Criteria - Detailed criteria for top 10")
print("  4. All Criteria - Complete list of 65 criteria")
print("  5. Grouping Impact - Before/after comparison")
print("  6. Summary - Key statistics")
