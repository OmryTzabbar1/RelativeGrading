#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract data from PDFs (first 2 pages only) and create individual Excel files.
Reads PDFs directly using PyMuPDF and extracts GitHub URLs from hyperlinks.
"""

import sys
import codecs
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import fitz  # PyMuPDF
except ImportError:
    print("ERROR: PyMuPDF not installed. Run: pip install PyMuPDF")
    sys.exit(1)

# Set UTF-8 encoding for stdout on Windows
if sys.platform == 'win32':
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')


def extract_github_url_from_pdf(pdf_path):
    """Extract GitHub URL from PDF hyperlinks and text (first 2 pages only)."""
    try:
        doc = fitz.open(pdf_path)
        max_pages = min(2, len(doc))

        for page_num in range(max_pages):
            page = doc[page_num]

            # First check hyperlinks
            links = page.get_links()
            for link in links:
                if 'uri' in link:
                    uri = link['uri']
                    if 'github.com' in uri:
                        doc.close()
                        return uri

            # Also search in text
            text = page.get_text()
            github_pattern = r'https?://github\.com/[^\s\)\]>]+'
            match = re.search(github_pattern, text)
            if match:
                doc.close()
                return match.group(0)

        doc.close()
        return ""
    except Exception as e:
        print(f"  ⚠ Error extracting GitHub URL: {e}")
        return ""


def extract_info_from_pdf(pdf_path):
    """
    Extract required information from PDF (first 2 pages only).
    Returns dict with: group_code, student1, student2, github, grade
    """
    info = {
        'group_code': '',
        'student1': '',
        'student2': '',
        'github': '',
        'grade': ''
    }

    try:
        doc = fitz.open(pdf_path)
        max_pages = min(2, len(doc))
        text = ""

        # Read first 2 pages
        for page_num in range(max_pages):
            page = doc[page_num]
            text += page.get_text()

        doc.close()

        # Try Format 1: "Team: Name1 (ID1), Name2 (ID2) | Repository:"
        team_pattern = r'Team:\s*([^|]+?)\s*\|'
        team_match = re.search(team_pattern, text)

        if team_match:
            team_text = team_match.group(1).strip()
            students = [s.strip() for s in team_text.split(',')]

            if len(students) >= 1:
                info['student1'] = students[0]
            if len(students) >= 2:
                info['student2'] = students[1]

            # Use first student's name as group code
            if info['student1']:
                name_match = re.search(r'^([^(]+)', info['student1'])
                if name_match:
                    info['group_code'] = name_match.group(1).strip().replace(' ', '_').lower()

        else:
            # Try Format 2: Hebrew format
            # Look for lines with: ID followed by Hebrew text
            # Pattern: 9-digit ID followed by Hebrew name
            lines = text.split('\n')

            # First, find the group name (usually appears before IDs, often in English)
            group_name = ""
            for i, line in enumerate(lines):
                line = line.strip()
                # Look for a line that looks like a group name (alphanumeric, underscores)
                if re.match(r'^[A-Za-z][A-Za-z0-9_\s-]+$', line) and len(line) < 50:
                    # Check if next lines have IDs (to confirm this is the group name)
                    if i + 1 < len(lines):
                        next_line = lines[i + 1].strip()
                        if re.search(r'\d{9}', next_line):
                            group_name = line
                            info['group_code'] = line.replace(' ', '_').lower()
                            break

            # Extract student lines (ID + Name in Hebrew)
            # Pattern: \d{9} followed by Hebrew characters
            student_pattern = r'(\d{9})\s+([\u0590-\u05FF\s]+)'
            student_matches = re.findall(student_pattern, text)

            if len(student_matches) >= 1:
                student_id, student_name = student_matches[0]
                # Reverse Hebrew text for better readability (right-to-left)
                info['student1'] = f"{student_name.strip()} ({student_id})"

            if len(student_matches) >= 2:
                student_id, student_name = student_matches[1]
                info['student2'] = f"{student_name.strip()} ({student_id})"

        # Extract self score
        # Try multiple patterns
        score_patterns = [
            r'(?:Our Self Score|Self Score):\s*(\d+)\s*/\s*100',  # English format
            r'(?:ימצע\s*ןויצ|ציון\s*עצמי)\s*[–-]\s*(\d+)',  # Hebrew format
            r'(?:ציון|הציון)\s*[:\s]*(\d+)',  # General Hebrew format
        ]

        for pattern in score_patterns:
            score_match = re.search(pattern, text, re.IGNORECASE)
            if score_match:
                info['grade'] = score_match.group(1)
                break

        # Extract GitHub URL (including from hyperlinks)
        info['github'] = extract_github_url_from_pdf(pdf_path)

    except Exception as e:
        print(f"  ⚠ Error reading PDF: {e}")

    return info


def create_student_excel(folder_path, participant_id, group_code, student1, student2, github, grade, pdf_filename):
    """Create an individual Excel file for a student submission."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission Info"

    # Headers
    headers = ["Field", "Value"]

    # Write headers with formatting
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    data = [
        ("Participant ID", participant_id),
        ("Group Code", group_code),
        ("Student 1", student1),
        ("Student 2", student2),
        ("GitHub Repository", github),
        ("Suggested Grade", grade),
        ("PDF Filename", pdf_filename)
    ]

    for row_idx, (field, value) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        cell = ws.cell(row=row_idx, column=2, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Make GitHub link clickable
        if field == "GitHub Repository" and value and value.startswith('http'):
            cell.hyperlink = value
            cell.font = Font(color="0563C1", underline="single")

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    # Save the Excel file in the student's folder
    output_path = folder_path / "submission_info.xlsx"
    wb.save(output_path)
    return output_path


def main():
    """Process all student submissions in WorkSubmissions05."""
    base_dir = Path(__file__).parent

    # Find all participant folders
    participant_folders = sorted([d for d in base_dir.iterdir()
                                  if d.is_dir() and d.name.startswith('Participant_')])

    print(f"Found {len(participant_folders)} student submissions\n")

    success_count = 0
    error_count = 0

    for folder in participant_folders:
        # Extract participant ID
        participant_id = folder.name.split('_')[1]

        # Find PDF file
        pdf_files = list(folder.glob('*.pdf'))

        if not pdf_files:
            print(f"⚠ No PDF found in {folder.name}")
            error_count += 1
            continue

        pdf_file = pdf_files[0]

        try:
            print(f"Processing {participant_id}... ", end='')

            # Extract info from PDF
            info = extract_info_from_pdf(str(pdf_file))

            # Create Excel file
            output_path = create_student_excel(
                folder,
                participant_id,
                info['group_code'],
                info['student1'],
                info['student2'],
                info['github'],
                info['grade'],
                pdf_file.name
            )

            print(f"✓")
            print(f"  Student 1: {info['student1']}")
            print(f"  Student 2: {info['student2']}")
            print(f"  GitHub: {info['github'] or 'Not found'}")
            print(f"  Self Score: {info['grade']}")
            print(f"  Excel: {output_path.name}\n")

            success_count += 1

        except Exception as e:
            print(f"✗ Error: {e}\n")
            error_count += 1

    print(f"\n{'='*60}")
    print(f"Processing complete!")
    print(f"  ✓ Success: {success_count}")
    print(f"  ✗ Errors: {error_count}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
